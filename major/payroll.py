# =================================
# 🏢 Enterprise Payroll Engine (Clean + Excel Export)
# =================================
import streamlit as st
import pandas as pd
from datetime import datetime
from core.database import supabase
import io
from core.database import save_data_saas, get_cached_data
import uuid
import streamlit as st
import pandas as pd
import uuid
from datetime import datetime
from io import BytesIO

def delete_data_saas(table_name, filters):
    """
    Deletes a record from Supabase based on a filter (e.g., payroll_id).
    """
    try:
        # Assuming 'supabase' is your initialized client in database.py
        response = supabase.table(table_name).delete().match(filters).execute()
        return True
    except Exception as e:
        st.error(f"Database Error: {e}")
        return False

def export_styled_excel(df, company="ZOE CONSULTS SMC LTD"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"

    # -----------------------------
    # Styles
    # -----------------------------
    blue = PatternFill("solid", fgColor="4A90E2")
    white_font = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right")

    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # -----------------------------
    # Title
    # -----------------------------
    ws.merge_cells("A1:W1")
    ws["A1"] = f"{datetime.now().strftime('%B %Y').upper()} PAYROLL ({company})"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    # -----------------------------
    # Header Row 1
    # -----------------------------
    ws.append([
        "S/NO","Employee Name","TIN","Designation","Mob No",
        "Account No","NSSF No",
        "Salary","Basic",
        "NO PAY","",
        "LST",
        "Gross Salary",
        "Deductions",
        "If yes calculate","","",
        "Other",
        "Total Deductions",
        "Nett Pay",
        "Total Tax",
        "10% NSSF",
        "15% NSSF"
    ])

    ws.merge_cells("J2:K2")
    ws.merge_cells("O2:Q2")

    # -----------------------------
    # Header Row 2
    # -----------------------------
    ws.append([
        "No","","","","","","",
        "ARREARS","Salary",
        "DAYS","Absenteeism",
        "Deductions",
        "",
        "P.A.Y.E","N.S.S.F","S.DRS/ADV",
        "Deduction",
        "",
        "",
        "",
        "",
        ""
    ])

    # Style headers
    for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=23):
        for cell in row:
            cell.fill = blue
            cell.font = white_font
            cell.alignment = center
            cell.border = thin

    # -----------------------------
    # Data
    # -----------------------------
    for i, r in df.iterrows():
        ws.append([
            i+1,
            r["employee"],
            r["tin"],
            r["designation"],
            r["mob_no"],
            r["account_no"],
            r["nssf_no"],
            r["arrears"],
            r["basic_salary"],
            0,
            r["absent_deduction"],
            r["lst"],
            r["gross_salary"],
            r["paye"],
            r["paye"],
            r["nssf_5"],
            r["advance_drs"],
            r["other_deductions"],
            r["paye"] + r["nssf_5"] + r["advance_drs"] + r["other_deductions"],
            r["net_pay"],
            r["paye"],
            r["nssf_10"],
            r["nssf_15"]
        ])

    # -----------------------------
    # Number formatting (DATA)
    # -----------------------------
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=8, max_col=23):
        for cell in row:
            cell.number_format = '#,##0'
            cell.alignment = right
            cell.border = thin

    # -----------------------------
    # Totals
    # -----------------------------
    total_row = ws.max_row + 1
    ws[f"A{total_row}"] = "TOTAL"

    for col in range(8, 24):
        letter = get_column_letter(col)
        ws[f"{letter}{total_row}"] = f"=SUM({letter}4:{letter}{total_row-1})"

    # Style totals row
    for col in range(1, 24):
        cell = ws[f"{get_column_letter(col)}{total_row}"]
        cell.font = bold
        cell.fill = blue
        cell.border = thin

    for col in range(8, 24):
        cell = ws[f"{get_column_letter(col)}{total_row}"]
        cell.number_format = '#,##0'
        cell.alignment = right

    # -----------------------------
    # Column widths
    # -----------------------------
    widths = [6,22,15,25,15,20,20] + [12]*16
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # -----------------------------
    # Freeze header
    # -----------------------------
    ws.freeze_panes = "A4"

    # -----------------------------
    # Save to memory
    # -----------------------------
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer

# ---------------------------------
# Payroll Calculation
# ---------------------------------
def run_manual_sync_calculations(basic, arrears, absent_deduct, advance, other):
    # 1. Gross Calculation
    gross = (float(basic) + float(arrears)) - float(absent_deduct)
    
    # 2. Local Service Tax (LST) Logic
    lst = 100000 / 12 if gross > 1000000 else 0
    
    # 3. NSSF Logic (Calculated but NOT subtracted from tax base)
    n5 = gross * 0.05
    n10 = gross * 0.10
    n15 = n5 + n10
    
    # 4. --- THE EXCEL MATCHING PAYE LOGIC ---
    # Based on your sheet: Tax = 25,000 + (30% * (Gross - 410,000))
    paye = 0
    if gross > 410000:
        excess = gross - 410000
        paye = 25000 + (0.30 * excess)
    elif gross > 235000:
        # Lower tier fallback
        paye = (gross - 235000) * 0.10
        
    # 5. Final Deductions & Net Pay
    # Deductions = PAYE + LST + NSSF(5%) + Advance + Other
    total_deductions = paye + lst + n5 + float(advance) + float(other)
    net = gross - total_deductions
    
    return {
        "gross": round(gross), "lst": round(lst), "n5": round(n5), 
        "n10": round(n10), "n15": round(n15), "paye": round(paye), "net": round(net)
    }

    # -----------------------------
    # NET PAY
    # -----------------------------
    net = gross - total_deductions

    return {
        "gross": gross,
        "lst": lst,
        "n5": nssf_5,
        "n10": nssf_10,
        "n15": nssf_15,
        "paye": paye,
        "net": round(net)
    }

# ---------------------------------
# Format Display Table
# ---------------------------------
def format_payroll_display(df):
    if df.empty:
        return df

    df = df.copy()

    df["NO"] = range(1, len(df) + 1)
    df["Salary"] = df["gross_salary"]
    df["Basic"] = df["basic_salary"]
    df["NO PAY DAYS"] = 0
    df["Absenteeism"] = df["absent_deduction"]

    df["Gross Salary"] = df["gross_salary"]
    df["Deductions"] = df["paye"]

    df["P.A.Y.E"] = df["paye"]
    df["N.S.S.F"] = df["nssf_5"]
    df["S.DRS/ADV"] = df["advance_drs"]
    df["Other"] = df["other_deductions"]

    df["Total Deductions"] = (
        df["paye"] + df["nssf_5"] + df["advance_drs"] + df["other_deductions"]
    )

    df["Nett Pay"] = df["net_pay"]
    df["Total Tax on Salary"] = df["paye"]
    df["10% NSSF"] = df["nssf_10"]
    df["15% NSSF"] = df["nssf_15"]

    return df[
        [
            "NO",
            "employee",
            "tin",
            "designation",
            "mob_no",
            "account_no",
            "nssf_no",
            "Salary",
            "Basic",
            "NO PAY DAYS",
            "Absenteeism",
            "Gross Salary",
            "Deductions",
            "P.A.Y.E",
            "N.S.S.F",
            "S.DRS/ADV",
            "Other",
            "Total Deductions",
            "Nett Pay",
            "Total Tax on Salary",
            "10% NSSF",
            "15% NSSF",
        ]
    ]

# ---------------------------------
# MAIN FUNCTION
# ---------------------------------
def show_payroll():

    # ==============================
    # 🎨 1. MASTER BUTTON STYLING (GLOBAL OVERRIDE)
    # ==============================
    st.markdown("""
    <style>
    /* MASTER BUTTON SELECTOR - Uniform look across the entire application workspace */
    div.stButton > button,
    div.stFormSubmitButton > button {
        background: linear-gradient(135deg, #1d4ed8 0%, #1e3a8a 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 12px !important;
        padding: 0.7rem 1.5rem !important;
        font-weight: 700 !important;
        font-size: 15px !important;
        height: 48px !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        gap: 8px !important;
        transition: all 0.25s ease-out !important;
        box-shadow: 0 4px 14px rgba(30, 58, 138, 0.2) !important;
        width: auto;
    }
    
    div.stButton > button[width="100%"] {
        width: 100% !important;
    }

    /* HOVER STATE (FLOAT EFFECT) */
    div.stButton > button:hover,
    div.stFormSubmitButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 25px rgba(30, 58, 138, 0.35) !important;
        background: linear-gradient(135deg, #2563eb 0%, #1d4ed8 100%) !important;
    }
    
    /* ACTIVE/CLICK STATE */
    div.stButton > button:active,
    div.stFormSubmitButton > button:active {
        transform: translateY(1px) !important;
        box-shadow: 0 2px 8px rgba(30, 58, 138, 0.15) !important;
    }

    /* Specific icon adjustments */
    .save-icon { font-size: 1.1em; color: #a3e635; margin-right: 2px; }
    .cancel-icon { font-size: 1.1em; color: #f87171; margin-right: 2px; }
    .download-icon { font-size: 1.1em; color: #60a5fa; margin-right: 2px; }
    </style>
    """, unsafe_allow_html=True)

    tenant = st.session_state.get("tenant_id")
    role = st.session_state.get("role")

    if not tenant or role != "Admin":
        st.error("🔒 Restricted: Only Admins only")
        return

    st.markdown("<h2 style='color:#4A90E2;'>🧾 Payroll</h2>", unsafe_allow_html=True)

    # Load Data
    payroll_df = get_cached_data("payroll")

    if payroll_df is not None and not payroll_df.empty:
        payroll_df.columns = payroll_df.columns.astype(str).str.strip().str.replace(" ", "_")
        payroll_df = payroll_df[payroll_df["tenant_id"].astype(str) == str(tenant)]
    else:
        payroll_df = pd.DataFrame()

    # Employee List
    employee_list = []
    if not payroll_df.empty and "employee" in payroll_df.columns:
        employee_list = sorted(payroll_df["employee"].dropna().astype(str).unique())

    # Tabs
    tab_process, tab_history = st.tabs(["💳 Process Payroll", "📜 Payroll History"])

    # =================================
    # PROCESS TAB
    # =================================
    with tab_process:
        with st.form("payroll_form", clear_on_submit=True):
    
            st.subheader("👤 Employee Info")
    
            selected_emp = st.selectbox("Select Employee", employee_list) if employee_list else None
            new_emp = st.text_input("Or Enter New Employee")
    
            employee_name = new_emp if new_emp else selected_emp
    
            c1, c2, c3 = st.columns(3)
            f_tin = c1.text_input("TIN")
            f_desig = c2.text_input("Designation")
            f_mob = c3.text_input("Mobile")
    
            c4, c5 = st.columns(2)
            f_acc = c4.text_input("Account No")
            f_nssf = c5.text_input("NSSF No")
    
            st.subheader("💰 Earnings & Deductions")
    
            # Added the LST toggle as requested to make it optional
            f_apply_lst = st.checkbox("Deduct Local Service Tax (LST)?", value=True)
    
            c6, c7, c8 = st.columns(3)
            f_arrears = c6.number_input("Arrears", min_value=0.0)
            f_basic = c7.number_input("Basic Salary", min_value=0.0)
            f_absent = c8.number_input("Absent Deduction", min_value=0.0)
    
            c9, c10 = st.columns(2)
            f_adv = c9.number_input("Advance", min_value=0.0)
            f_other = c10.number_input("Other Deductions", min_value=0.0)
    
            # Form submission buttons accept string labels and are styled globally by the CSS wrapper
            if st.form_submit_button("💳 Save Payroll", use_container_width=True):
    
                if not employee_name or f_basic <= 0:
                    st.error("Enter valid employee & salary")
                    return
    
                month_str = datetime.now().strftime("%Y-%m")
    
                if not payroll_df.empty:
                    duplicate = payroll_df[
                        (payroll_df["employee"] == employee_name) &
                        (payroll_df["month"] == month_str)
                    ]
                    if not duplicate.empty:
                        st.warning("Payroll already exists for this month")
                        return
    
                # Updated to pass the f_apply_lst toggle to your compute_payroll function
                calc = compute_payroll(f_basic, f_arrears, f_absent, f_adv, f_other, apply_lst=f_apply_lst)
    
                new_row = pd.DataFrame([{
                    "payroll_id": str(uuid.uuid4()),
                    "employee": employee_name,
                    "tin": f_tin,
                    "designation": f_desig,
                    "mob_no": f_mob,
                    "account_no": f_acc,
                    "nssf_no": f_nssf,
                    "arrears": f_arrears,
                    "basic_salary": f_basic,
                    "absent_deduction": f_absent,
                    "gross_salary": calc["gross"],
                    "lst": calc["lst"],
                    "paye": calc["paye"],
                    "nssf_5": calc["n5"],
                    "nssf_10": calc["n10"],
                    "nssf_15": calc["n15"],
                    "advance_drs": f_adv,
                    "other_deductions": f_other,
                    "net_pay": calc["net"],
                    "date": datetime.utcnow().isoformat(),
                    "month": month_str,
                    "tenant_id": str(tenant)
                }])
    
                if save_data_saas("payroll", new_row):
                    get_cached_data.clear()
                    st.success(f"✅ Saved for {employee_name}")
                    st.rerun()
                    
    # =================================
    # HISTORY TAB
    # =================================
    with tab_history:

        if payroll_df.empty:
            st.info("No payroll records")
            return

        st.markdown("### 📊 Payroll Sheet")

        # Display Dataframe
        display_df = format_payroll_display(payroll_df)
        try:
            formatted_df = format_with_commas(display_df)
        except NameError:
            formatted_df = display_df 

        st.dataframe(formatted_df, use_container_width=True)

        # -----------------------------
        # 🛠️ EDIT / DELETE SECTION
        # -----------------------------
        st.markdown("---")
        st.subheader("🛠️ Manage Records")
        
        # Select record by Employee and date for clarity
        record_options = payroll_df.apply(lambda x: f"{x['employee']} ({x['month']}) | ID: {x['payroll_id'][:8]}", axis=1).tolist()
        selected_record_str = st.selectbox("Select a record to Edit or Delete", options=record_options)

        if selected_record_str:
            # Extract the actual payroll_id from the selection string
            sel_id = selected_record_str.split("| ID: ")[1].strip()
            # Match back to the full ID in the dataframe
            full_record = payroll_df[payroll_df['payroll_id'].str.contains(sel_id)].iloc[0]

            col_edit, col_del = st.columns(2)

            with col_edit:
                if st.button("📝 Edit Selected Record", use_container_width=True, key=f"edit_btn_{sel_id}"):
                    st.warning("To edit: Adjust details in the 'Process Payroll' tab with the same name and month to overwrite, or use the database editor.")
            
            with col_del:
                if st.button("🗑️ Delete Record", use_container_width=True, key=f"del_btn_{sel_id}", type="primary"):
                    # This will now find the function in your imported modules
                    if delete_data_saas("payroll", {"payroll_id": full_record['payroll_id']}):
                        get_cached_data.clear()
                        st.success(f"Deleted payroll for {full_record['employee']}")
                        st.rerun()

        st.markdown("---")
        # -----------------------------
        # 📄 DOWNLOADS
        # -----------------------------
        c1, c2 = st.columns(2)
        with c1:
            csv = payroll_df.to_csv(index=False).encode("utf-8")
            st.download_button(
                label="📄 Download CSV",
                data=csv,
                file_name=f"Payroll_{datetime.now().strftime('%Y%m%d')}.csv",
                use_container_width=True,
                key="payroll_csv_dl"
            )

        with c2:
            excel_file = export_styled_excel(payroll_df)
            st.download_button(
                label="📥 Download Styled Excel",
                data=excel_file,
                file_name=f"Payroll_Styled_{datetime.now().strftime('%B_%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="payroll_xlsx_dl"
            )
