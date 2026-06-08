import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from supabase import create_client, Client
import io
import base64
import json
import os
import re
from datetime import datetime, timedelta
from fpdf import FPDF
from streamlit_calendar import calendar
import bcrypt
from twilio.rest import Client as TwilioClient
import time
import uuid
import extra_streamlit_components as stx



# --- Page Config stays here ---
st.set_page_config(
    page_title="Peak-Lenders Africa",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)
# Constants
SESSION_TIMEOUT = 30

# ============================================================
# 🔌 1. SUPABASE INIT (ROBUST & GLOBAL)
# ============================================================
@st.cache_resource
def init_supabase():
    try:
        url = st.secrets.get("SUPABASE_URL") or os.getenv("SUPABASE_URL")
        key = st.secrets.get("SUPABASE_KEY") or os.getenv("SUPABASE_KEY")
        if not url or not key:
            return None
        return create_client(url, key)
    except Exception:
        return None

supabase = init_supabase()

if supabase is None:
    st.warning("⚠️ Supabase credentials not configured. Please check your secrets.")

# ============================================================
# 🏢 2. CONFIGURATION & SESSION STATE INITIALIZATION
# ============================================================
SESSION_TIMEOUT = 15  # Minutes
MAX_ATTEMPTS = 5
LOCKOUT_MINUTES = 10

if "tenant_id" not in st.session_state:
    st.session_state["tenant_id"] = None
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False
if "theme_color" not in st.session_state:
    st.session_state["theme_color"] = "#1E3A8A" 
if "data_version" not in st.session_state:
    st.session_state["data_version"] = 0
if "view" not in st.session_state:
    st.session_state["view"] = "login"

# 🔒 Hydrate active connection if session cookies exist
if "auth_session" in st.session_state and supabase:
    try:
        supabase.auth.set_session(
            st.session_state["auth_session"].access_token,
            st.session_state["auth_session"].refresh_token
        )
    except Exception:
        pass

# ============================================================
# ⚡ 3. SECURITY & TENANT CORE UTILITIES
# ============================================================
def get_tenant_id():
    return st.session_state.get("tenant_id")

def require_tenant():
    if not st.session_state.get("tenant_id"):
        st.error("Session expired or unauthorized access. Please log in again.")
        st.stop()

def check_session_timeout():
    if not st.session_state.get("logged_in"):
        return

    last_activity = st.session_state.get("last_activity", datetime.now())
    idle_time = (datetime.now() - last_activity).total_seconds() / 60

    if idle_time > SESSION_TIMEOUT:
        auth_keys = ["logged_in", "authenticated", "user_id", "tenant_id", "company", "auth_session", "role"]
        for key in auth_keys:
            st.session_state.pop(key, None)
        st.session_state["view"] = "login"
        st.warning("Session timed out due to inactivity. Please log in again.")
        st.rerun()

    st.session_state["last_activity"] = datetime.now()

# ============================================================
# 📊 4. CACHED DATA LAYER ENGINE
# ============================================================
@st.cache_data(ttl=600, show_spinner=False)
def get_cached_data(table_name):
    try:
        if supabase is None:
            return pd.DataFrame()

        require_tenant()
        tenant_id = get_tenant_id()

        res = supabase.table(table_name).select("*").eq("tenant_id", tenant_id).execute()

        if res.data:
            df = pd.DataFrame(res.data)
            df.columns = df.columns.str.strip().str.lower()
            return df
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Database Fetch Error [{table_name}]: {e}")
        return pd.DataFrame()

def save_data(table_name, dataframe):
    try:
        if supabase is None:
            st.error("❌ Database not connected")
            return False

        require_tenant()

        if dataframe is None or dataframe.empty:
            st.error("No Data to commit.")
            return False

        df_to_save = dataframe.copy()
        df_to_save["tenant_id"] = get_tenant_id()
        records = df_to_save.replace({np.nan: None}).to_dict("records")

        response = supabase.table(table_name).upsert(records).execute()
        if response.data:
            st.success(f"Saved {len(response.data)} record(s)")
            return True
        return False
    except Exception as e:
        st.error(f"DB Error [{table_name}]: {e}")
        return False

# ============================================================
# 📤 5. STORAGE & AUDIT HELPERS
# ============================================================
def safe_audit_log(payload):
    try:
        if supabase:
            supabase.table("audit_logs").insert(payload).execute()
    except Exception:
        pass

def upload_image(file, bucket="collateral-photos"):
    try:
        if supabase is None:
            st.error("Storage unavailable: Supabase client not initialized.")
            return None

        require_tenant()
        tenant_id = get_tenant_id()

        clean_name = re.sub(r'[^a-zA-Z0-9._-]', '_', file.name)
        file_path = f"{tenant_id}/{datetime.now().strftime('%Y%m%d_%H%M%S')}_{clean_name}"

        supabase.storage.from_(bucket).upload(
            path=file_path,
            file=file.getvalue(),
            file_options={"content-type": file.type}
        )

        return supabase.storage.from_(bucket).get_public_url(file_path)
    except Exception as e:
        st.error(f"Upload failed: {e}")
        return None

# ============================================================
# 🎨 6. THEME CUSTOMIZATION ENGINE
# ============================================================
def apply_master_theme():
    # 🎨 Read the brand color actively set via the settings panel color picker
    brand_color = st.session_state.get("theme_color", "#2B3F87")
    
    # Helper: Convert clean Hex strings into functional translucent RGB arrays for glass layouts
    hex_color = brand_color.lstrip('#')
    try:
        rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    except Exception:
        rgb = (43, 63, 135) # Fallback to core Slate Blue state

    st.markdown(f"""
    <style>
    /* 1. Base Workspace Canvas Background */
    .stApp {{
        background: linear-gradient(135deg, #F8FAFC 0%, #F1F5F9 100%) !important;
    }}

    /* 2. DYNAMIC SIDEBAR TINT WITH GLASSMORPHISM ENGINE */
    /* Mixes 85% pure white with 15% of your customized settings brand color tint */
    [data-testid="stSidebar"] {{
        background-color: rgba({rgb[0]}, {rgb[1]}, {rgb[2]}, 0.08) !important;
        background-image: linear-gradient(
            180deg, 
            rgba(255, 255, 255, 0.65) 0%, 
            rgba(255, 255, 255, 0.45) 100%
        ) !important;
        backdrop-filter: blur(25px) saturate(180%) !important;
        -webkit-backdrop-filter: blur(25px) saturate(180%) !important;
        border-right: 1px solid rgba({rgb[0]}, {rgb[1]}, {rgb[2]}, 0.12) !important;
        box-shadow: 4px 0 30px 0 rgba({rgb[0]}, {rgb[1]}, {rgb[2]}, 0.05) !important;
    }}
    
    [data-testid="stSidebar"] > div:first-child {{
        background: transparent !important;
    }}

    /* 3. Typography Realignment with High-Contrast Slate Elements */
    [data-testid="stSidebar"] .stMarkdown p, 
    [data-testid="stSidebar"] label, 
    [data-testid="stSidebar"] span,
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3,
    [data-testid="stSidebar"] [data-testid="stWidgetLabel"] p {{
        color: #0F172A !important;
        font-weight: 500 !important;
    }}

    [data-testid="stSidebar"] p[data-testid="stCaptionText"],
    [data-testid="stSidebar"] .stMarkdown caption {{
        color: #475569 !important;
    }}

    /* 4. Selectbox Input Structural Modifications */
    div[data-baseweb="select"] > div {{
        background: rgba(255, 255, 255, 0.85) !important;
        border: 1px solid rgba({rgb[0]}, {rgb[1]}, {rgb[2]}, 0.15) !important;
        border-radius: 12px !important;
    }}
    
    div[data-baseweb="select"] span {{
        color: #0F172A !important;
    }}

    /* 5. Navigation Items Radio Group Styling */
    [data-testid="stSidebar"] [data-testid="stRadio"] div[role="radiogroup"] {{
        background: rgba(255, 255, 255, 0.4) !important;
        border-radius: 16px !important;
        padding: 6px !important;
        border: 1px solid rgba({rgb[0]}, {rgb[1]}, {rgb[2]}, 0.06) !important;
    }}

    [data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"] {{
        background: transparent;
        transition: all 0.2s ease-in-out;
        padding: 8px 12px !important;
        border-radius: 10px !important;
        margin-bottom: 3px !important;
        width: 100% !important;
    }}

    [data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"]:hover {{
        background: rgba(255, 255, 255, 0.75) !important;
    }}

    /* 6. Action Button Syncing (Matches your dynamic hex code perfectly) */
    [data-testid="stSidebar"] button {{
        background: rgba(255, 255, 255, 0.75) !important;
        color: #0F172A !important;
        backdrop-filter: blur(8px) !important;
        border: 1px solid rgba({rgb[0]}, {rgb[1]}, {rgb[2]}, 0.15) !important;
        border-radius: 12px !important;
        font-weight: 500 !important;
        transition: all 0.2s ease !important;
    }}

    [data-testid="stSidebar"] button:hover {{
        background: #FFFFFF !important;
        border-color: {brand_color} !important;
        color: {brand_color} !important;
        transform: translateY(-0.5px);
    }}
    </style>
    """, unsafe_allow_html=True)

def auth_styles():
    st.markdown("""
    <style>
    .stApp { background: linear-gradient(135deg,#EEF4FF,#F8FAFC); }
    .stButton > button { border-radius: 12px; height: 42px; font-weight: 600; }
    .stFormSubmitButton > button { background: linear-gradient(90deg,#1D4ED8,#2563EB); color: white; }
    </style>
    """, unsafe_allow_html=True)

# ============================================================
# 🏢 7. DIALOGS & USER ROUTING INTERFACES
# ============================================================
@st.dialog("🏢 Register Company")
def admin_company_registration():
    st.caption("Create your organization account")
    with st.form("company_reg_form"):
        company_name = st.text_input("Organization Name", placeholder="Peak-Lenders Africa")
        admin_name = st.text_input("Admin Full Name", placeholder="John Doe")
        email = st.text_input("Business Email", placeholder="admin@company.com")
        pwd = st.text_input("Password", type="password")
        submit = st.form_submit_button("✨ Create Organization", use_container_width=True)

    if submit:
        if not all([company_name, admin_name, email, pwd]):
            st.error("All fields are required")
            return
        try:
            res = supabase.auth.sign_up({"email": email, "password": pwd})
            if not res.user:
                st.error("Registration failed")
                return

            tenant_id = str(uuid.uuid4())
            company_code = f"{company_name[:3].upper()}{uuid.uuid4().int % 999}"

            supabase.table("tenants").insert({"id": tenant_id, "name": company_name, "company_code": company_code}).execute()
            supabase.table("users").insert({"id": res.user.id, "name": admin_name, "email": email, "tenant_id": tenant_id, "role": "Admin"}).execute()

            safe_audit_log({
                "user_id": res.user.id, "action": "CREATE_COMPANY", "tenant_id": tenant_id, "timestamp": datetime.now().isoformat()
            })
            st.success(f"✅ Organization created!\n\nCompany Code: **{company_code}**")
        except Exception as e:
            st.error(f"Registration failed: {e}")

@st.dialog("👥 Staff Registration")
def view_staff_signup():
    st.caption("Create a staff account")
    with st.form("staff_signup_form"):
        company_code_input = st.text_input("Company Registration Code", placeholder="PEA123")
        name = st.text_input("Full Name", placeholder="Jane Doe")
        email = st.text_input("Email Address", placeholder="staff@company.com")
        pwd = st.text_input("Password", type="password")
        submit = st.form_submit_button("🚀 Create Staff Account", use_container_width=True)

    if submit:
        if not all([company_code_input, name, email, pwd]):
            st.error("All fields are required")
            return
        try:
            tenant_query = supabase.table("tenants").select("*").eq("company_code", company_code_input.strip().upper()).execute()
            if not tenant_query.data:
                st.error("Invalid Company Registration Code.")
                return

            tenant = tenant_query.data[0]
            res = supabase.auth.sign_up({"email": email, "password": pwd})
            if not res.user:
                st.error("Signup failed")
                return

            supabase.table("users").insert({"id": res.user.id, "name": name, "email": email, "tenant_id": tenant["id"], "role": "Staff"}).execute()
            st.success("✅ Staff account created successfully!")
        except Exception as e:
            st.error(f"Signup failed: {e}")

@st.dialog("🔑 Forgot Password")
def forgot_password_page():
    st.caption("Reset your account password")
    email = st.text_input("Registered Email", placeholder="you@example.com")
    if st.button("📩 Send Reset Link", use_container_width=True):
        if not email:
            st.error("Email required")
            return
        try:
            supabase.auth.reset_password_for_email(email)
            st.success("✅ Password reset email sent")
        except Exception as e:
            st.error(f"Reset failed: {e}")

# ============================================================
# 🔐 8. MAIN ACCESS PORTAL (FIXED INVERSION SCOPE)
# ============================================================
def login_page():
    auth_styles()
    st.write(" ")
    
    _, center, _ = st.columns([1, 1.4, 1])
    with center:
        with st.container(border=True):
            st.markdown("## 💰 PEAK-LENDERS AFRICA")
            st.caption("Secure Business Login Portal")
            st.divider()

            with st.form("login_form"):
                # 🏢 Switched from Company Code back to Business Name
                company_name_input = st.text_input(
                    "Business Name", 
                    placeholder="e.g., Peak-Lenders Africa"
                ).strip()
                
                email = st.text_input("Email Address", placeholder="Enter email").strip().lower()
                pwd = st.text_input("Password", type="password", placeholder="Enter password")
                submit = st.form_submit_button("🔓 Access Dashboard", use_container_width=True)

            if submit:
                if not all([company_name_input, email, pwd]):
                    st.error("All input fields are required.")
                else:
                    try:
                        # Step 1: Attempt normal Supabase authentication
                        res = supabase.auth.sign_in_with_password({"email": email, "password": pwd})
                        if not res or not res.user:
                            st.error("Invalid credentials")
                            return

                        # Step 2: Fetch the user profile along with the linked tenant data
                        user_query = supabase.table("users").select("*, tenants(*)").eq("id", res.user.id).execute()
                        if not user_query.data:
                            st.error("Profile matching registration missing.")
                            return

                        user_profile = user_query.data[0]
                        tenant_data = user_profile.get("tenants") or {}
                        
                        # Extract the exact company name from the DB
                        db_company_name = tenant_data.get("name", "").strip()

                        # Step 3: Case-insensitive match on the clean Business Name string
                        if db_company_name.lower() != company_name_input.lower():
                            st.error(f"Access Denied: This account is not verified under '{company_name_input}'.")
                            return

                        # Step 4: Establish verified session states
                        st.session_state["logged_in"] = True
                        st.session_state["authenticated"] = True
                        st.session_state["user_id"] = user_profile["id"]
                        st.session_state["tenant_id"] = user_profile["tenant_id"]
                        st.session_state["role"] = user_profile.get("role", "Staff")
                        st.session_state["company"] = db_company_name
                        st.session_state["last_activity"] = datetime.now()
                        st.session_state["view"] = "main"

                        st.success("Login authorized!")
                        time.sleep(0.4)
                        st.rerun()

                    except Exception as e:
                        st.error(f"Login failed: {e}")

            st.write(" ")
            col1, col2, col3 = st.columns(3)
            with col1:
                if st.button("🏢 Register", use_container_width=True):
                    admin_company_registration()
            with col2:
                if st.button("👥 Staff", use_container_width=True):
                    view_staff_signup()
            with col3:
                if st.button("🔑 Reset", use_container_width=True):
                    forgot_password_page()
# ============================================================
# 🌐 9. APPLICATION ROUTER GATEWAY
# ============================================================
def run_auth_ui(supabase_client): # Add the parameter here
    check_session_timeout()
    
    view = st.session_state.get("view", "login")
    
    if st.session_state.get("logged_in") and view != "main":
        st.session_state["view"] = "main"
        st.rerun()

    if view == "login":
        login_page() # If login_page accesses global client
   
# ============================================================
# ⚡ ENTERPRISE SIDEBAR (FAST + CLEAN + SAFE)
# ============================================================

import streamlit as st
import time
import pandas as pd


# ============================================================
# 🧠 CACHE: TENANTS (CRITICAL SPEED FIX)
# ============================================================
@st.cache_data(ttl=600, show_spinner=False)
def get_tenants():
    try:
        res = supabase.table("tenants")\
            .select("id, name, brand_color, logo_url")\
            .execute()
        return res.data or []
    except:
        return []


# ============================================================
# 🖼️ CACHE: LOGO BUILDER
# ============================================================
@st.cache_data(ttl=600)
def build_logo_url(logo_val):
    if not logo_val:
        return None

    if str(logo_val).startswith("http"):
        return logo_val

    base = st.secrets.get("SUPABASE_URL", "").strip("/")
    if not base:
        return None

    return f"{base}/storage/v1/object/public/company-logos/{logo_val}"


# ============================================================
# 🚦 SIDEBAR (ENTERPRISE VERSION)
# ============================================================
def render_sidebar():

    # --------------------------------------------------------
    # 1. FETCH TENANTS (CACHED)
    # --------------------------------------------------------
    tenants = get_tenants()
    tenant_map = {t["name"]: t for t in tenants}

    selected_page = "Overview"

    with st.sidebar:
        st.markdown("")

        # ----------------------------------------------------
        # 2. BUSINESS SELECTOR
        # ----------------------------------------------------
        if not tenant_map:
            st.warning("No businesses found")
            st.stop()

        options = list(tenant_map.keys())
        current_tenant_id = st.session_state.get("tenant_id")

        default_index = 0
        for i, name in enumerate(options):
            if tenant_map[name]["id"] == current_tenant_id:
                default_index = i
                break

        selected_name = st.selectbox(
            "🏢 Business Portal",
            options,
            index=default_index,
            key="sidebar_business_selector"
        )

        active_company = tenant_map[selected_name]
        new_tenant_id = active_company.get("id")

        # 🔥 CRITICAL FIX: If the tenant changed, sync IDs and flush cached calculations
        if current_tenant_id != new_tenant_id:
            st.session_state["tenant_id"] = new_tenant_id
            st.session_state["company"] = selected_name
            # Clear data cache pools so the next page render is forced to read fresh DB values
            if hasattr(st, "cache_data"):
                st.cache_data.clear()
            st.rerun()

        # ----------------------------------------------------
        # 3. UPDATE THEME (FAST STATE ONLY)
        # ----------------------------------------------------
        st.session_state["theme_color"] = active_company.get(
            "brand_color",
            "#1E3A8A"
        )

        # ----------------------------------------------------
        # 4. LOGO + BRANDING
        # ----------------------------------------------------
        logo_url = build_logo_url(active_company.get("logo_url"))

        if logo_url:
            # Removed border-radius: 50% to show your natural wide logo cleanly
            st.markdown(
                f"""
                <div style="display:flex; justify-content:center; margin-bottom:10px;">
                    <img src="{logo_url}?t={int(time.time())}"
                        style="max-width:180px; max-height:80px; object-fit:contain;" />
                </div>
                """,
                unsafe_allow_html=True
            )
        else:
            st.markdown("<h3 style='text-align:center;'>🏢</h3>", unsafe_allow_html=True)

        # Added dynamic classes matching our slate glass theme rules
        st.markdown(
            f"""
            <div class="sidebar-company-title" style='text-align:center; font-weight:600; font-size:16px; margin-top:5px;'>
                {selected_name}
            </div>
            <div class="sidebar-company-sub" style='text-align:center; font-size:11px; letter-spacing:1px; margin-bottom:10px;'>
                FINANCE CORE
            </div>
            """,
            unsafe_allow_html=True
        )

        st.divider()

        # ----------------------------------------------------
        # 5. NAVIGATION MENU (FAST STATE ONLY)
        # ----------------------------------------------------
        menu = {
            "Overview": "📈",
            "loans": "💵",
            "borrowers": "👥",
            "Collateral": "🛡️",
            "Calendar": "📅",
            "Ledger": "📄",
            "Payroll": "💳",
            "Expenses": "📉",
            "Overdue Tracker": "🚨",
            "Payments": "💰",
            "Reports": "📊",
            "Settings": "⚙️"
        }

        menu_options = [f"{v} {k}" for k, v in menu.items()]
        current_page = st.session_state.get("current_page", "Overview")

        try:
            default_ix = list(menu.keys()).index(current_page)
        except:
            default_ix = 0

        selection = st.radio(
            "Navigation",
            menu_options,
            index=default_ix,
            label_visibility="collapsed",
            key="nav_radio"
        )

        selected_page = selection.split(" ", 1)[1]
        st.session_state["current_page"] = selected_page

        st.divider()

        # ----------------------------------------------------
        # 6. LOGOUT (SAFE + CLEAN)
        # ----------------------------------------------------
        if st.session_state.get("logged_in"):
            if st.button("🚪 Logout", use_container_width=True):
                # Clear session safely
                for k in list(st.session_state.keys()):
                    del st.session_state[k]

                st.session_state["logged_in"] = False
                st.session_state["view"] = "login"
                st.rerun()

    return selected_page

# ==========================================
# 1. CORE PAGE FUNCTIONS & LAYOUT
# ==========================================

st.set_page_config(layout="wide", page_title="Financial Control Center", page_icon="🏛️")

if "auto_refresh_tick" not in st.session_state:
    st.session_state.auto_refresh_tick = 0

def soft_refresh():
    st.session_state.auto_refresh_tick += 1

def get_Active_color():
    return st.session_state.get("theme_color", "#1E3A8A")

@st.cache_data(ttl=60, show_spinner=False)
def load_cached(name):
    try:
        # Mocking implementation layer for execution safety
        return get_cached_data(name)
    except:
        return pd.DataFrame()

# =========================================================
# HELPERS
# =========================================================

def normalize(df):
    try:
        if df is None or not isinstance(df, pd.DataFrame) or df.empty:
            return pd.DataFrame()
        df = df.copy()
        df.columns = (
            df.columns.astype(str)
            .str.strip()
            .str.lower()
            .str.replace(" ", "_")
        )
        return df
    except:
        return pd.DataFrame()

def safe_numeric(df, cols):
    try:
        if df is None or df.empty:
            return pd.Series(dtype="float64")
        for c in cols:
            if c in df.columns:
                return pd.to_numeric(df[c], errors="coerce").fillna(0)
        return pd.Series(0.0, index=df.index)
    except:
        return pd.Series(0.0)

def safe_date(df, cols):
    try:
        if df is None or df.empty:
            return pd.Series(dtype="datetime64[ns]")
        for c in cols:
            if c in df.columns:
                return pd.to_datetime(df[c], errors="coerce")
        return pd.Series(pd.NaT, index=df.index)
    except:
        return pd.Series(pd.NaT)

def first_existing(df, cols):
    try:
        for c in cols:
            if c in df.columns:
                return c
        return None
    except:
        return None

# =========================================================
# MAIN DASHBOARD VIEW
# =========================================================
def show_dashboard_view():
    brand_color = get_Active_color()

    try:
        # --- GLOBAL MODERN INJECTED CSS ---
        st.markdown(f"""
        <style>
        /* Modern Glassmorphic Dashboard Header */
        .dashboard-header {{
            background: linear-gradient(135deg, {brand_color}, #0F172A);
            padding: 30px;
            border-radius: 16px;
            margin-bottom: 30px;
            color: white;
            box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.1), 0 8px 10px -6px rgba(0, 0, 0, 0.1);
        }}
        
        /* Unified Metric Boxes */
        .metric-box {{
            padding: 20px;
            border-radius: 14px;
            color: #FFFFFF;
            box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05);
            transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1);
            border: 1px solid rgba(255,255,255,0.05);
        }}
        .metric-box:hover {{
            transform: translateY(-4px);
            box-shadow: 0 12px 20px -3px rgba(0,0,0,0.12);
        }}
        .metric-title {{
            font-size: 11px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 1.2px;
            opacity: 0.8;
            margin-bottom: 8px;
        }}
        .metric-value {{
            font-size: 28px;
            font-weight: 700;
            letter-spacing: -0.5px;
            line-height: 1.2;
        }}
        .metric-sub {{
            font-size: 12px;
            opacity: 0.75;
            margin-top: 6px;
        }}
        
        /* Premium Gradient Assignments */
        .card-blue {{ background: linear-gradient(135deg, #2563EB, #1D4ED8); }}
        .card-green {{ background: linear-gradient(135deg, #10B981, #047857); }}
        .card-red {{ background: linear-gradient(135deg, #EF4444, #B91C1C); }}
        .card-slate {{ background: linear-gradient(135deg, #64748B, #475569); }}
        
        /* Glassmorphic mini cards for light/dark layout engine */
        .mini-kpi {{
            padding: 16px;
            border-radius: 12px;
            background: rgba(155, 155, 155, 0.08);
            border: 1px solid rgba(155, 155, 155, 0.15);
            text-align: center;
        }}
        </style>
        """, unsafe_allow_html=True)

        # --- UI HEADER ---
        st.markdown(f"""
        <div class="dashboard-header">
            <h1 style='margin:0; font-size:32px; font-weight:700; letter-spacing:-0.5px;'>🏛️ Financial Control Center</h1>
            <p style='margin:6px 0 0 0; opacity:0.85; font-size:15px;'>Real-time tracking engine for loans, expenses, and asset portfolio performance.</p>
        </div>
        """, unsafe_allow_html=True)

        # --- DATA INGESTION ---
        loans_df = normalize(load_cached("loans"))
        payments_df = normalize(load_cached("payments"))
        expenses_df = normalize(load_cached("expenses"))
        borrowers_df = normalize(load_cached("borrowers"))
        
        if loans_df.empty:
            st.info("👋 Welcome! No active data layers found. Populate underlying tables to visualize tracking pipelines.")
            return
        
        # --- CLEAN COLUMNS SAFELY ---
        required_loan_cols = [
            "status", "principal", "amount", "interest", "interest_amount",
            "balance", "total_repayable", "amount_paid", "paid", "end_date", "due_date"
        ]
        for col in required_loan_cols:
            if col not in loans_df.columns:
                loans_df[col] = 0
                
        loans_df["status"] = loans_df["status"].astype(str).str.upper().str.strip()
        loans_df.loc[loans_df["status"] == "0", "status"] = "ACTIVE"

        # --- ENGINE PIPELINE CONFIGURATION ---
        loans_df["principal_n"] = safe_numeric(loans_df, ["principal", "amount"])
        loans_df["interest_n"] = safe_numeric(loans_df, ["interest", "interest_amount"])
        total_repayable = safe_numeric(loans_df, ["balance", "total_repayable"])
        amount_paid = safe_numeric(loans_df, ["amount_paid", "paid"])
        
        loans_df["balance_n"] = (total_repayable - amount_paid).clip(lower=0)
        
        # --- EXPENSES CONFIGURATION ---
        if expenses_df.empty:
            total_expenses = 0.0
        else:
            expenses_df["amount"] = safe_numeric(expenses_df, ["amount"])
            total_expenses = float(expenses_df["amount"].sum())
        
        # --- OVERDUE CALCULATIONS ---
        today = pd.Timestamp.now().normalize()
        loans_df["due_date_dt"] = safe_date(loans_df, ["end_date", "due_date"])
        
        overdue_mask = (
            loans_df["due_date_dt"].notna() & 
            (loans_df["due_date_dt"] < today) & 
            (~loans_df["status"].isin(["CLEARED", "PAID"]))
        )
        overdue_count = int(overdue_mask.sum())
        
        # --- PORTFOLIO EXPOSURE LOGIC ---
        if "cycle_no" not in loans_df.columns:
            loans_df["cycle_no"] = 1
        loans_df["cycle_no"] = pd.to_numeric(loans_df["cycle_no"], errors="coerce").fillna(1)
        
        original_loans = loans_df[loans_df["cycle_no"] == 1]
        total_principal = float(original_loans["principal_n"].sum())
        total_interest = float(loans_df["interest_n"].sum())
        
        # --- ACTION ALERTS ---
        if overdue_count >= 5:
            st.toast(f"⚠️ Action Required: {overdue_count} Overdue Loans Detected.", icon="⚠️")

        # --- DISPLAY MAIN HIGH-TIER METRICS ---
        def render_metric_card(container, title, value, subtitle, css_class):
            container.markdown(f"""
            <div class="metric-box {css_class}">
                <div class="metric-title">{title}</div>
                <div class="metric-value">{value}</div>
                <div class="metric-sub">{subtitle}</div>
            </div>
            """, unsafe_allow_html=True)
        
        m1, m2, m3, m4 = st.columns(4)
        render_metric_card(m1, "Active Principal", f"UGX {total_principal:,.0f}", "Portfolio Asset Value", "card-blue")
        render_metric_card(m2, "Interest Income", f"UGX {total_interest:,.0f}", "Expected Yield Gross", "card-green")
        render_metric_card(m3, "Operational Costs", f"UGX {total_expenses:,.0f}", "Total Outgoings Logged", "card-red")
        render_metric_card(m4, "Critical Alerts", str(overdue_count), "Overdue Portfolio Loans", "card-slate")
        
        st.write("---")
        
        # ==========================================
        # VISUALIZATION PIPELINES (ROW 1)
        # ==========================================
        col_l, col_r = st.columns([2, 1])
        
        with col_l:
            st.markdown("### 📈 Revenue Margins vs Expenses")
            if not payments_df.empty:
                try:
                    pay_date_col = first_existing(payments_df, ["date", "payment_date", "created_at"])
                    pay_amt_col = first_existing(payments_df, ["amount", "paid", "payment"])
                    
                    if pay_date_col and pay_amt_col:
                        rev_df = payments_df.copy()
                        rev_df["date_dt"] = pd.to_datetime(rev_df[pay_date_col], errors="coerce")
                        rev_df["amount_n"] = pd.to_numeric(rev_df[pay_amt_col], errors="coerce").fillna(0)
                        rev_df = rev_df.dropna(subset=["date_dt"])
                        rev_df["month"] = rev_df["date_dt"].dt.to_period("M").dt.to_timestamp()
                        
                        monthly_rev = rev_df.groupby("month", as_index=False)["amount_n"].sum().rename(columns={"amount_n": "Revenue"})
                        
                        if not expenses_df.empty:
                            exp_df = expenses_df.copy()
                            exp_date_col = first_existing(exp_df, ["payment_date", "date", "created_at"])
                            exp_df["date_dt"] = pd.to_datetime(exp_df[exp_date_col], errors="coerce")
                            exp_df["amount_n"] = pd.to_numeric(exp_df["amount"], errors="coerce").fillna(0)
                            exp_df = exp_df.dropna(subset=["date_dt"])
                            exp_df["month"] = exp_df["date_dt"].dt.to_period("M").dt.to_timestamp()
                            monthly_exp = exp_df.groupby("month", as_index=False)["amount_n"].sum().rename(columns={"amount_n": "Expenses"})
                        else:
                            monthly_exp = pd.DataFrame(columns=["month", "Expenses"])
                            
                        trend_df = pd.merge(monthly_rev, monthly_exp, on="month", how="outer").fillna(0).sort_values("month")
                        
                        fig = px.line(
                            trend_df, x="month", y=["Revenue", "Expenses"],
                            template="plotly_white",
                            color_discrete_map={"Revenue": "#10B981", "Expenses": "#EF4444"}
                        )
                        fig.update_traces(mode="lines+markers", line=dict(width=3), marker=dict(size=7))
                        fig.update_layout(
                            height=320, margin=dict(l=10, r=10, t=15, b=10),
                            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                            xaxis_title="", yaxis_title="UGX Denominated", hovermode="x unified"
                        )
                        st.plotly_chart(fig, use_container_width=True)
                except Exception as e:
                    st.caption(f"Trend engine calculation bypass: {e}")
            else:
                st.info("Performance monitoring requires transaction inputs.")
                
        with col_r:
            st.markdown("### 🎯 Portfolio Segmentation")
            if not loans_df.empty:
                try:
                    clean_status = loans_df["status"].replace({
                        "CURRENT": "ACTIVE", "ONGOING": "ACTIVE", 
                        "COMPLETE": "PAID", "CLOSED": "PAID", 
                        "LATE PAYMENT": "LATE", "DEFAULTED": "DEFAULT"
                    })
                    status_data = clean_status.value_counts().reset_index()
                    status_data.columns = ["status", "count"]
                    
                    fig_pie = px.pie(
                        status_data, names="status", values="count", hole=0.7,
                        color="status", color_discrete_map={"ACTIVE": "#2563EB", "PAID": "#10B981", "LATE": "#F59E0B", "DEFAULT": "#EF4444"}
                    )
                    fig_pie.update_traces(textposition="outside", textinfo="percent")
                    fig_pie.update_layout(
                        height=320, margin=dict(l=10, r=10, t=15, b=10),
                        showlegend=True, legend=dict(orientation="h", y=-0.1),
                        annotations=[dict(text=f"<b>{len(loans_df)}</b><br>Loans", x=0.5, y=0.5, font_size=18, showarrow=False)]
                    )
                    st.plotly_chart(fig_pie, use_container_width=True)
                except Exception as e:
                    st.caption(f"Segmentation pipeline failure: {e}")
            else:
                st.info("No distribution values found.")

        # ==========================================
        # CHANNELS & HISTORICAL LOGIC (ROW 2)
        # ==========================================
        st.write("---")
        t1, t2 = st.columns(2)
        
        with t1:
            st.markdown("### 📊 Issuance Timeline Performance")
            try:
                graph_df = loans_df.copy()
                graph_df["date_dt"] = safe_date(graph_df, ["start_date", "created_at"])
                graph_df = graph_df.dropna(subset=["date_dt"])
                
                if not graph_df.empty:
                    graph_df["month"] = graph_df["date_dt"].dt.to_period("M").dt.to_timestamp()
                    timeline_df = graph_df.groupby("month")[["principal_n", "interest_n"]].sum().reset_index().sort_values("month")
                    timeline_df.rename(columns={"principal_n": "Capital Deployed", "interest_n": "Expected Yield"}, inplace=True)
                    
                    fig_portfolio = px.bar(
                        timeline_df, x="month", y=["Capital Deployed", "Expected Yield"],
                        barmode="group", template="plotly_white",
                        color_discrete_map={"Capital Deployed": brand_color, "Expected Yield": "#34D399"}
                    )
                    fig_portfolio.update_layout(
                        height=350, hovermode="x unified", xaxis_title="",
                        yaxis_title="UGX", legend=dict(orientation="h", y=1.05)
                    )
                    st.plotly_chart(fig_portfolio, use_container_width=True)
                else:
                    st.info("Insufficent asset tracking historical context.")
            except Exception as e:
                st.caption(f"Asset charting pipeline fallback: {e}")
                
        with t2:
            st.markdown("### 💸 Live Outgoings Registry")
            if not expenses_df.empty:
                try:
                    df = expenses_df.copy()
                    df["amount"] = safe_numeric(df, ["amount"])
                    df["date"] = safe_date(df, ["date", "payment_date"])
                    df = df.sort_values("date", ascending=False)
                    latest = df.head(5)
                    
                    # Mini Unified Transparent KPI Engine Block
                    k1, k2, k3 = st.columns(3)
                    k1.markdown(f'<div class="mini-kpi"><span style="font-size:12px;opacity:0.7;">Top 5 Total</span><br><b style="font-size:16px;color:#EF4444;">UGX {latest["amount"].sum():,.0f}</b></div>', unsafe_allow_html=True)
                    k2.markdown(f'<div class="mini-kpi"><span style="font-size:12px;opacity:0.7;">Batch Mean</span><br><b style="font-size:16px;">UGX {latest["amount"].mean():,.0f}</b></div>', unsafe_allow_html=True)
                    k3.markdown(f'<div class="mini-kpi"><span style="font-size:12px;opacity:0.7;">Entries Ledger</span><br><b style="font-size:16px;">{len(latest)}</b></div>', unsafe_allow_html=True)
                    
                    st.write("")
                    
                    # Safe Presentation Mapping Layer
                    display_df = latest.copy()
                    display_df["category"] = display_df.get("category", "General").fillna("General")
                    if "date" in display_df.columns and not display_df["date"].isna().all():
                        display_df["date"] = display_df["date"].dt.strftime("%Y-%m-%d")
                    else:
                        display_df["date"] = "N/A"
                        
                    final_df = display_df[["category", "amount", "date"]]
                    
                    # Safe formatting execution natively
                    st.dataframe(
                        final_df.style.format({"amount": "UGX {:,.0f}"}),
                        use_container_width=True, hide_index=True
                    )
                except Exception as e:
                    st.error(f"Render Registry Fault: {e}")
            else:
                st.info("System accounts are clear. No operational costs logged.")

        # ==========================================
        # SYSTEM FILE EXPORT PIPELINES
        # ==========================================
        st.write("---")
        c1, c2 = st.columns(2)
        
        with c1:
            st.download_button(
                label="📥 Export Capital Ledger (CSV)",
                data=loans_df.to_csv(index=False).encode("utf-8"),
                file_name=f"capital_portfolio_{pd.Timestamp.now().strftime('%Y-%m-%d')}.csv",
                mime="text/csv", use_container_width=True
            )
        with c2:
            st.download_button(
                label="⬇️ Export Expense Ledger (CSV)",
                data=expenses_df.to_csv(index=False).encode("utf-8"),
                file_name=f"expense_portfolio_{pd.Timestamp.now().strftime('%Y-%m-%d')}.csv",
                mime="text/csv", use_container_width=True
            )

    except Exception as e:
        st.error(f"Critical Dashboard Core Recovery Engaged: {str(e)}")



import streamlit as st
import pandas as pd
import uuid

# ==============================
# 🚀 BORROWERS ENGINE (PRODUCTION)
# ==============================

def show_borrowers():

    # ==============================
    # 🎨 BRANDING & THEME
    # ==============================
    brand_color = st.session_state.get("theme_color", "#1E3A8A")
    st.markdown(f"<h2 style='color:{brand_color};'>🚀 Borrowers Registry</h2>", unsafe_allow_html=True)

    # ==============================
    # 🔐 TENANT SESSION CHECK
    # ==============================
    tenant_id = st.session_state.get("tenant_id")
    if not tenant_id:
        st.error("Session expired. Please log in again.")
        st.stop()

    # ==============================
    # 🧠 SAFE HELPERS (INTERNAL)
    # ==============================
    def safe_df(df):
        """Ensure the object is a valid pandas DataFrame."""
        return df if isinstance(df, pd.DataFrame) else pd.DataFrame()

    def safe_numeric(df, col, default=0.0, as_int=False):
        """Safely extract a numeric column from a DataFrame."""
        if not isinstance(df, pd.DataFrame) or df.empty:
            return pd.Series(dtype="int64" if as_int else "float64")

        if col in df.columns:
            s = pd.to_numeric(df[col], errors="coerce")
        else:
            s = pd.Series([default] * len(df), index=df.index)

        s = s.fillna(default)
        return s.astype("int64") if as_int else s

    # ==============================
    # 📥 LOAD & NORMALIZE DATA
    # ==============================
    borrowers_df = safe_df(get_cached_data("borrowers"))
    loans_df = safe_df(get_cached_data("loans"))
    
    # Force lowercase column names for consistency
    for df in [borrowers_df, loans_df]:
        if not df.empty:
            df.columns = df.columns.str.strip().str.lower()

    # 🔐 STRICT TENANT FILTERS (Prevents Data Leakage Across Companies)
    if not borrowers_df.empty and "tenant_id" in borrowers_df.columns:
        borrowers_df = borrowers_df[borrowers_df["tenant_id"].astype(str) == str(tenant_id)]

    if not loans_df.empty and "tenant_id" in loans_df.columns:
        loans_df = loans_df[loans_df["tenant_id"].astype(str) == str(tenant_id)]

    # Ensure required structural columns exist
    required_cols = ["id", "name", "phone", "email", "status", "national_id", "next_of_kin", "address"]
    for col in required_cols:
        if col not in borrowers_df.columns:
            borrowers_df[col] = ""

    # ==============================
    # 🔗 DATA LINKAGE
    # ==============================
    if not borrowers_df.empty and not loans_df.empty:
        borrowers_df["id"] = borrowers_df["id"].astype(str).str.strip()
        loans_df["borrower_id"] = loans_df["borrower_id"].astype(str).str.strip()
        
        bor_map = dict(zip(borrowers_df["id"], borrowers_df["name"]))
        loans_df["borrower"] = loans_df["borrower_id"].map(bor_map).fillna("Unknown Borrower")
    else:
        loans_df["borrower"] = "Unknown"

    # ==============================
    # 🔥 REAL-TIME RISK ENGINE
    # ==============================
    risk_map = {}
    if not loans_df.empty:
        loans_df["balance"] = safe_numeric(loans_df, "balance")
        date_col = "end_date" if "end_date" in loans_df.columns else "due_date"
        loans_df["parsed_due_date"] = pd.to_datetime(loans_df[date_col], errors="coerce")

        today = pd.Timestamp.today().normalize()
        loans_df["days_overdue"] = (today - loans_df["parsed_due_date"]).dt.days
        loans_df["days_overdue"] = loans_df["days_overdue"].apply(lambda x: x if x > 0 else 0)
        loans_df["is_overdue"] = (loans_df["days_overdue"] > 0) & (loans_df["balance"] > 0)

        risk_df = loans_df.groupby("borrower_id").agg({
            "balance": "sum",
            "is_overdue": "sum",
            "days_overdue": "max"
        }).reset_index()

        risk_df.rename(columns={"balance": "exposure", "is_overdue": "overdue_count", "days_overdue": "max_days"}, inplace=True)

        def classify_risk(row):
            if row["overdue_count"] == 0: return "🟢 Healthy"
            elif row["max_days"] <= 7: return "🟡 Watch"
            elif row["max_days"] <= 30: return "🟠 Risk"
            else: return "🔴 Critical"

        risk_df["risk_label"] = risk_df.apply(classify_risk, axis=1)
        risk_map = risk_df.set_index("borrower_id").to_dict("index")

    # ==============================
    # 📑 UI NAVIGATION
    # ==============================
    tab_view, tab_add = st.tabs(["📋 View Borrowers", "➕ Add Borrower"])

    with tab_add:
        with st.form("add_borrower_form", clear_on_submit=True):
            st.markdown(f"<h4 style='color: {brand_color};'>📝 Register New Borrower</h4>", unsafe_allow_html=True)
            c1, c2 = st.columns(2)
            name = c1.text_input("Full Name*")
            phone = c2.text_input("Phone Number*")
            email = c1.text_input("Email Address")
            nid = c2.text_input("National ID / NIN")
            addr = c1.text_input("Physical Address")
            nok = c2.text_input("Next of Kin (Name & Contact)")
            
            if st.form_submit_button("🚀 Save Borrower Profile", use_container_width=True):
                if name and phone:
                    new_id = str(uuid.uuid4())
                    new_entry = pd.DataFrame([{
                        "id": new_id, "name": name, "phone": phone, "email": email,
                        "national_id": nid, "address": addr, "next_of_kin": nok,
                        "status": "Active", "tenant_id": str(tenant_id)
                    }])
                
                    if save_data_saas("borrowers", new_entry):
                        st.success(f"✅ {name} registered successfully!")
                        st.cache_data.clear()
                        st.rerun()
                else:
                    st.error("⚠️ Full Name and Phone Number are required.")

    with tab_view:
        st.markdown("### 👥 Borrowers Registry")
        search = st.text_input("🔍 Search by name or phone...").lower()
    
        if not borrowers_df.empty:
            df = borrowers_df.copy()
    
            for col in ["name", "phone", "national_id", "next_of_kin", "status"]:
                df[col] = df[col].astype(str)
    
            def get_risk_label(b_id):
                r = risk_map.get(str(b_id), {})
                return r.get("risk_label", "🟢 Healthy")
    
            df["Risk Status"] = df["id"].apply(get_risk_label)
    
            # Search logic
            df_filtered = df[
                df["name"].str.lower().str.contains(search, na=False) |
                df["phone"].str.contains(search, na=False)
            ]
    
            if not df_filtered.empty:
                def style_risk(val):
                    if "🔴" in val: return "color: #EF4444; font-weight:700;"
                    elif "🟠" in val: return "color: #F97316; font-weight:700;"
                    elif "🟡" in val: return "color: #F59E0B; font-weight:700;"
                    else: return "color: #10B981; font-weight:700;"
    
                display_df = df_filtered[["name", "phone", "national_id", "next_of_kin", "Risk Status", "status"]].copy()
                display_df.columns = ["Borrower Name", "Phone", "National ID", "Next of Kin", "Risk Status", "Status"]
                display_df["Status"] = display_df["Status"].str.upper()
    
                styled_df = display_df.style.map(style_risk, subset=["Risk Status"])
    
                st.dataframe(styled_df, use_container_width=True, hide_index=True)
    
                # --- Interactive Dropdown Selection ---
                st.markdown("### 🎯 Management Actions")
                
                # Derive current index safely to keep selectbox synchronized
                current_selection = st.session_state.get("selected_borrower")
                default_index = 0
                borrower_list = df_filtered["name"].tolist()
                
                if current_selection:
                    matched_rows = df_filtered[df_filtered["id"] == current_selection]
                    if not matched_rows.empty:
                        default_index = borrower_list.index(matched_rows.iloc[0]["name"]) + 1

                selected_name = st.selectbox(
                    "Select borrower:",
                    ["-- Choose borrower --"] + borrower_list,
                    index=default_index
                )
    
                if selected_name != "-- Choose borrower --":
                    sel_id = df_filtered[df_filtered["name"] == selected_name]["id"].values[0]
                    if st.session_state.get("selected_borrower") != sel_id:
                        st.session_state["selected_borrower"] = sel_id
                        st.rerun()
            else:
                st.info("No records match your search criteria.")
        else:
            st.info("The registry is currently empty.")

    # ==============================
    # 👤 BORROWER PROFILE PANEL
    # ==============================
    selected_id = st.session_state.get("selected_borrower")

    if selected_id:
        st.write("---")
        st.markdown(f"### 👤 Profile Detail: {str(selected_id)[:8]}")

        borrower_query = borrowers_df[borrowers_df["id"].astype(str) == str(selected_id)]

        if not borrower_query.empty:
            borrower = borrower_query.iloc[0]

            with st.container(border=True):
                c1, c2 = st.columns(2)
                upd_name = c1.text_input("Name", borrower["name"])
                upd_phone = c2.text_input("Phone", borrower["phone"])
                upd_email = c1.text_input("Email", borrower["email"])
                upd_nid = c2.text_input("National ID", borrower["national_id"])
                
                c3, c4 = st.columns(2)
                upd_nok = c3.text_input("Next of Kin", borrower["next_of_kin"])
                upd_addr = c4.text_input("Address", borrower["address"])

                # 📊 NESTED LOAN HISTORY
                st.markdown("#### 💳 Loan Statement")
                user_loans = loans_df[loans_df["borrower_id"].astype(str) == str(selected_id)].copy()

                if not user_loans.empty:
                    # Fix applied: capital 'D' on DateColumn
                    st.dataframe(
                        user_loans, 
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "id": None, "tenant_id": None, "borrower_id": None, "borrower": None,
                            "principal": st.column_config.NumberColumn("Principal", format="%d UGX"),
                            "interest": st.column_config.NumberColumn("Interest", format="%d UGX"),
                            "balance": st.column_config.NumberColumn("Balance", format="%d UGX"),
                            "total_repayable": st.column_config.NumberColumn("Total Due", format="%d UGX"),
                            "start_date": st.column_config.DateColumn("Date Issued"),
                            "end_date": st.column_config.DateColumn("Due Date"),
                        }
                    )
                    
                    csv = user_loans.to_csv(index=False).encode('utf-8')
                    st.download_button(
                        label="📥 Download Statement (CSV)",
                        data=csv,
                        file_name=f"Statement_{upd_name.replace(' ', '_')}.csv",
                        mime="text/csv",
                    )
                else:
                    st.info("This borrower has no loan history.")

                # 🛠️ ACTION BUTTONS
                st.write("---")
                act_c1, act_c2, act_c3 = st.columns([1, 1, 2])

                if act_c1.button("💾 Save Changes", use_container_width=True):
                    # Explicit row modification updates safely
                    idx = borrowers_df.index[borrowers_df["id"].astype(str) == str(selected_id)].tolist()[0]
                    borrowers_df.at[idx, "name"] = upd_name
                    borrowers_df.at[idx, "phone"] = upd_phone
                    borrowers_df.at[idx, "email"] = upd_email
                    borrowers_df.at[idx, "national_id"] = upd_nid
                    borrowers_df.at[idx, "next_of_kin"] = upd_nok
                    borrowers_df.at[idx, "address"] = upd_addr
                    
                    if save_data_saas("borrowers", borrowers_df):
                        st.success("Profile Updated Successfully")
                        st.cache_data.clear()
                        st.rerun()

                if act_c2.button("🗑️ Delete", use_container_width=True):
                    updated_df = borrowers_df[borrowers_df["id"].astype(str) != str(selected_id)]
                    if save_data_saas("borrowers", updated_df):
                        st.warning("Profile Removed")
                        st.cache_data.clear()
                        st.session_state.pop("selected_borrower", None)
                        st.rerun()
                
                if act_c3.button("❌ Close Profile", use_container_width=True):
                    st.session_state.pop("selected_borrower", None)
                    st.rerun()

import streamlit as st
import pandas as pd
import uuid
from datetime import datetime, timedelta

# ==============================
# 🔐 SAAS TENANT CONTEXT (UUID SAFE)
# ==============================
def get_current_tenant():
    """Returns tenant UUID only"""
    tenant_id = st.session_state.get("tenant_id", None)
    if tenant_id in [None, "", "default_tenant"]:
        return None
    return str(tenant_id)


# ==============================
# 🧠 DATABASE ADAPTER (MULTI-TENANT SAFE)
# ==============================
def get_data(table_name):
    tenant_id = str(get_current_tenant()).strip()
    df = get_cached_data(table_name)

    if df is None:
        return pd.DataFrame()

    if df.empty:
        return df

    df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")

    if "tenant_id" in df.columns:
        df["tenant_id"] = df["tenant_id"].astype(str).str.strip()
        df = df[df["tenant_id"] == tenant_id].copy()

    return df.reset_index(drop=True)

def save_data_saas(table_name, df):
    tenant_id = get_current_tenant()
    if tenant_id:
        df["tenant_id"] = str(tenant_id)
    return save_data(table_name, df)


# ==============================
# 13. LOANS MANAGEMENT PAGE
# ==============================
def show_loans():
    st.markdown(
        "<h2 style='color: #0A192F;'>💵 Loans Management</h2>",
        unsafe_allow_html=True
    )

    # ------------------------------
    # LOAD DATA
    # ------------------------------
    loans_df = get_data("loans")
    borrowers_df = get_data("borrowers")
    payments_df = get_data("payments")

    # ------------------------------
    # SAFETY FALLBACKS
    # ------------------------------
    if loans_df is None or loans_df.empty:
        loans_df = pd.DataFrame(columns=[
            "id", "sn", "loan_id_label", "parent_loan_id", "borrower_id",
            "borrower", "loan_type", "principal", "interest", "total_repayable",
            "amount_paid", "balance", "status", "start_date", "end_date", "cycle_no", "tenant_id"
        ])

    if borrowers_df is None:
        borrowers_df = pd.DataFrame()

    if payments_df is None:
        payments_df = pd.DataFrame()

    # ------------------------------
    # REQUIRED DEFAULTS
    # ------------------------------
    required_defaults = {
        "id": "", "sn": "", "loan_id_label": "", "parent_loan_id": "", "borrower_id": "",
        "borrower": "", "loan_type": "", "principal": 0.0, "interest": 0.0,
        "total_repayable": 0.0, "amount_paid": 0.0, "balance": 0.0, "status": "ACTIVE",
        "start_date": "", "end_date": "", "cycle_no": 1, "tenant_id": ""
    }

    for col, val in required_defaults.items():
        if col not in loans_df.columns:
            loans_df[col] = val

    loans_df = loans_df.copy()
    payments_df = payments_df.copy()

    # ------------------------------
    # TYPE CLEANUP
    # ------------------------------
    loans_df["id"] = loans_df["id"].astype(str)
    loans_df["borrower_id"] = loans_df["borrower_id"].astype(str)
    loans_df["parent_loan_id"] = loans_df["parent_loan_id"].fillna("").astype(str)

    if not payments_df.empty and "loan_id" in payments_df.columns:
        payments_df["loan_id"] = payments_df["loan_id"].astype(str)

    # ------------------------------
    # NUMERIC CLEANUP
    # ------------------------------
    for col in ["principal", "interest", "total_repayable", "amount_paid", "balance"]:
        loans_df[col] = pd.to_numeric(loans_df[col], errors="coerce").fillna(0)

    if not payments_df.empty and "amount" in payments_df.columns:
        payments_df["amount"] = pd.to_numeric(payments_df["amount"], errors="coerce").fillna(0)

    # ------------------------------
    # DATE CLEANUP
    # ------------------------------
    for col in ["start_date", "end_date"]:
        loans_df[col] = pd.to_datetime(loans_df[col], errors="coerce")

    # ------------------------------
    # PAYMENT SYNC
    # ------------------------------
    loans_df["amount_paid"] = 0  # ensure column always exists
    if not payments_df.empty and "loan_id" in payments_df.columns:
        pay_sums = payments_df.groupby("loan_id")["amount"].sum()
        loans_df["amount_paid"] = loans_df["id"].map(pay_sums).fillna(0)

    loans_df["balance"] = (loans_df["total_repayable"] - loans_df["amount_paid"]).clip(lower=0)

    # ==============================
    # SERIAL ENGINE
    # ==============================
    loans_df["sn"] = loans_df["sn"].astype(str).str.strip()
    
    existing_nums = []
    for val in loans_df["sn"]:
        if val.startswith("LN-"):
            try:
                existing_nums.append(int(val.replace("LN-", "")))
            except ValueError:
                pass
    
    next_sn_val = max(existing_nums, default=0)
    
    id_to_sn = dict(zip(loans_df["id"], loans_df["sn"]))
    id_to_parent = dict(zip(loans_df["id"], loans_df["parent_loan_id"].fillna("").astype(str).str.strip()))
    
    # Lineage Walk Assignment
    for i in loans_df.index:
        current_id = loans_df.at[i, "id"]
        if id_to_sn.get(current_id, "").startswith("LN-"):
            continue
    
        visited = set()
        parent_id = id_to_parent.get(current_id, "")
        inherited_sn = ""
    
        while parent_id and parent_id not in ("nan", ""):
            if parent_id in visited:
                break
            visited.add(parent_id)
    
            parent_sn = id_to_sn.get(parent_id, "")
            if parent_sn.startswith("LN-"):
                inherited_sn = parent_sn
                break
            parent_id = id_to_parent.get(parent_id, "")
    
        if inherited_sn:
            loans_df.at[i, "sn"] = inherited_sn
            id_to_sn[current_id] = inherited_sn
        else:
            next_sn_val += 1
            new_sn = f"LN-{next_sn_val:04d}"
            loans_df.at[i, "sn"] = new_sn
            id_to_sn[current_id] = new_sn
    
    loans_df["sn"] = loans_df["sn"].str.upper()
    loans_df = loans_df.sort_values(by=["sn", "start_date", "id"])
    loans_df["cycle_no"] = loans_df.groupby("sn").cumcount() + 1
    loans_df = loans_df.sort_values(by=["sn", "cycle_no", "start_date"])
    loans_df["status"] = ""
    
    for sn_val, grp in loans_df.groupby("sn"):
        indices = grp.index.tolist()
        latest_idx = indices[-1]
    
        if len(indices) > 1:
            loans_df.loc[indices[:-1], "status"] = "BCF"
    
        latest_row = loans_df.loc[latest_idx]
        if abs(latest_row["balance"]) < 1.0:
            loans_df.at[latest_idx, "status"] = "CLEARED"
        else:
            if int(latest_row["cycle_no"]) == 1:
                bytes_stat = "ACTIVE"
            else:
                bytes_stat = "PENDING"
            loans_df.at[latest_idx, "status"] = bytes_stat
    
    # Final overrides
    mask_zero_balance = loans_df["balance"] <= 0
    mask_not_bcf = loans_df["status"] != "BCF"
    loans_df.loc[mask_zero_balance & mask_not_bcf, "status"] = "CLEARED"
    
    loans_df = loans_df.sort_values(by=["sn", "cycle_no"]).reset_index(drop=True)
    loans_df["loan_id_label"] = loans_df["sn"].str.replace("LN-", "", regex=False).str.zfill(4)

    # ==============================
    # 🔄 DATABASE SYNC ENGINE
    # ==============================
    raw_db_df = get_cached_data("loans")
    
    def needs_update(row):
        if raw_db_df is None or raw_db_df.empty:
            return True
        db_match = raw_db_df[raw_db_df["id"] == row["id"]]
        if db_match.empty:
            return True
        db_row = db_match.iloc[0]
        return (str(db_row.get("sn", "")) != str(row["sn"]) or 
                str(db_row.get("loan_id_label", "")) != str(row["loan_id_label"]) or
                int(db_row.get("cycle_no", 0)) != int(row["cycle_no"]))

    to_sync = loans_df[loans_df.apply(needs_update, axis=1)]
    
    if not to_sync.empty:
        with st.status("🔄 Syncing Serial Numbers to Database...", expanded=False) as status:
            for _, row in to_sync.iterrows():
                sync_data = {
                    "sn": row["sn"],
                    "loan_id_label": row["loan_id_label"],
                    "cycle_no": int(row["cycle_no"])
                }
                try:
                    supabase.table("loans").update(sync_data).eq("id", row["id"]).execute()
                except Exception as e:
                    st.error(f"Error syncing row {row['id']}: {e}")
            
            st.cache_data.clear()
            status.update(label="✅ Database Serial Numbers Synced!", state="complete", expanded=False)
            st.rerun()

    # ------------------------------
    # BORROWER MAP
    # ------------------------------
    if not borrowers_df.empty:
        borrowers_df["id"] = borrowers_df["id"].astype(str)
        bor_map = dict(zip(borrowers_df["id"], borrowers_df["name"]))
        loans_df["borrower"] = loans_df["borrower_id"].map(bor_map).fillna(loans_df["borrower"]).fillna("Unknown")

    if not borrowers_df.empty and "status" in borrowers_df.columns:
        Active_borrowers = borrowers_df[borrowers_df["status"].astype(str).str.upper() == "ACTIVE"]
    else:
        Active_borrowers = pd.DataFrame(columns=["id", "name"])

    # ==============================
    # TABS DESIGN INTERFACE
    # ==============================
    tab_view, tab_add, tab_manage, tab_actions = st.tabs([
        "📑 Portfolio View", "➕ New Loan", "🛠️ Manage/Edit", "⚙️ Actions"
    ])

    # ==============================
    # TAB VIEW
    # ==============================
    with tab_view:
        search_query = st.text_input("🔍 Search Loan / borrower", key="loan_search_main")

        # Create a local copy for filtering
        filtered_loans = loans_df.copy() if not loans_df.empty else pd.DataFrame()

        # 🎯 FIX 1: Compute Fiscal Year Fields BEFORE metric computations
        if not filtered_loans.empty:
            if "fiscal_year" not in filtered_loans.columns:
                start_dt = pd.to_datetime(filtered_loans.get("start_date"), errors="coerce")
                start_dt = start_dt.fillna(pd.to_datetime(filtered_loans.get("created_at", pd.Timestamp.today())))
                
                fiscal_years_list = []
                for dt in start_dt:
                    if dt.month >= 7:
                        fiscal_years_list.append(f"{dt.year}/{dt.year + 1}")
                    else:
                        fiscal_years_list.append(f"{dt.year - 1}/{dt.year}")
                
                filtered_loans["fiscal_year"] = fiscal_years_list
            
            fy_unique = sorted(filtered_loans["fiscal_year"].dropna().unique().tolist())
            fy_selected = st.selectbox("Filter by Fiscal Year", ["All"] + fy_unique)
            
            if fy_selected != "All":
                filtered_loans = filtered_loans[filtered_loans["fiscal_year"] == fy_selected]

        # Apply general manual keyword search queries
        if not filtered_loans.empty and search_query:
            filtered_loans = filtered_loans[
                filtered_loans.apply(lambda r: search_query.lower() in str(r).lower(), axis=1)
            ]

        # 📊 PORTFOLIO METRICS (Now perfectly handles selection changes)
        if not filtered_loans.empty:
            total_loans = filtered_loans["sn"].nunique()
            original_loans = filtered_loans[filtered_loans["cycle_no"] == 1]  
            total_principal = original_loans["principal"].sum()
            total_repayable = filtered_loans["total_repayable"].sum()
            total_paid = filtered_loans["amount_paid"].sum()
            total_pending = filtered_loans[filtered_loans["status"] == "PENDING"]["total_repayable"].sum()

            col1, col2, col3, col4 = st.columns(4)
            col1.markdown(f'<div style="background: linear-gradient(135deg, #3b82f6, #1e3a8a); padding:15px; border-radius:10px; color:white; text-align:center;"><div style="font-size:14px;">📄 Total Loans</div><div style="font-size:22px;font-weight:bold;">{total_loans}</div></div>', unsafe_allow_html=True)
            col2.markdown(f'<div style="background: linear-gradient(135deg, #10b981, #065f46); padding:15px; border-radius:10px; color:white; text-align:center;"><div style="font-size:14px;">💰 Principal</div><div style="font-size:22px;font-weight:bold;">{total_principal:,.0f}</div></div>', unsafe_allow_html=True)
            col3.markdown(f'<div style="background: linear-gradient(135deg, #f59e0b, #92400e); padding:15px; border-radius:10px; color:white; text-align:center;"><div style="font-size:14px;">💳 Paid</div><div style="font-size:22px;font-weight:bold;">{total_paid:,.0f}</div></div>', unsafe_allow_html=True)
            col4.markdown(f'<div style="background: linear-gradient(135deg, #ef4444, #991b1b); padding:15px; border-radius:10px; color:white; text-align:center;"><div style="font-size:14px;">⏳ Total Pending</div><div style="font-size:22px;font-weight:bold;">{total_pending:,.0f}</div></div>', unsafe_allow_html=True)

            st.markdown("---")

        # 📋 LOAN DATA TABLE
        if filtered_loans.empty:
            st.warning("No matching loans found.")
        else:
            filtered_loans["balance"] = filtered_loans["total_repayable"] - filtered_loans["amount_paid"]
        
            show_cols = ["sn", "loan_id_label", "borrower", "cycle_no", "principal", "total_repayable", "amount_paid", "balance", "start_date", "end_date", "status"]
            
            def style_entire_row(row):
                val = str(row["status"]).upper().strip()
                color_map = {
                    "ACTIVE": "background-color:#dbeafe;color:#1e40af;font-weight:bold;",
                    "PENDING": "background-color:#fee2e2;color:#991b1b;font-weight:bold;",
                    "CLEARED": "background-color:#d1fae5;color:#065f46;",
                    "BCF": "background-color:#ffedd5;color:#9a3412;",
                    "CLOSED": "background-color:#f3f4f6;color:#374151;"
                }
                return [color_map.get(val, "")] * len(row)
        
            styled_df = (
                filtered_loans[show_cols].style
                .apply(style_entire_row, axis=1)
                .format({
                    "principal": "{:,.0f}", "amount_paid": "{:,.0f}",
                    "total_repayable": "{:,.0f}", "balance": "{:,.0f}"
                })
            )
        
            st.dataframe(styled_df, column_order=show_cols, use_container_width=True, hide_index=True)
            
    # ==============================       
    # TAB ADD LOAN
    # ==============================
    with tab_add:
        if Active_borrowers.empty:
            st.info("💡 Tip: Activate borrower first.")
        else:
            with st.form("loan_issue_form_v2"):
                st.markdown("<h4 style='color:#0A192F;'>📝 Create New Loan Agreement</h4>", unsafe_allow_html=True)
                col1, col2 = st.columns(2)

                borrower_map = dict(zip(Active_borrowers["name"], Active_borrowers["id"]))
                selected_name = col1.selectbox("Select borrower", list(borrower_map.keys()))
                selected_id = str(borrower_map[selected_name]).strip()

                amount = col1.number_input("Principal amount (UGX)", min_value=0, step=50000)
                date_issued = col1.date_input("Start date", value=datetime.now())

                loan_type = col2.selectbox("Loan type", ["Business", "Personal", "Emergency", "Other"])
                interest_rate = col2.number_input("Monthly Interest Rate (%)", min_value=0.0, step=0.5)
                date_due = col2.date_input("Due date", value=date_issued + timedelta(days=30))

                total_due = amount + (amount * interest_rate / 100)
                st.info(f"Preview: Total Repayable {total_due:,.0f} UGX")

                submit = st.form_submit_button("🚀 Confirm & Issue Loan")

                if submit:
                    tenant_id = get_current_tenant()
                    if not tenant_id:
                        st.error("Tenant session missing.")
                        st.stop()

                    if selected_id == "":
                        st.error("borrower ID missing.")
                        st.stop()

                    loan_data = {
                        "id": str(uuid.uuid4()),
                        "sn": "PENDING",  # FIX 2: Set safe initialization value to prevent infinite engine sync rerun loops
                        "loan_id_label": "PENDING",
                        "parent_loan_id": None,
                        "borrower_id": selected_id,
                        "borrower": selected_name,
                        "loan_type": loan_type,
                        "principal": float(amount),
                        "interest": float(amount * interest_rate / 100),
                        "total_repayable": float(total_due),
                        "amount_paid": 0.0,
                        "balance": float(total_due),
                        "status": "ACTIVE",
                        "start_date": str(date_issued),
                        "end_date": str(date_due),
                        "cycle_no": 1,
                        "tenant_id": tenant_id
                    }

                    if save_data("loans", pd.DataFrame([loan_data])):
                        st.success("✅ Loan issued successfully.")
                        st.cache_data.clear()
                        st.session_state.pop("loans", None)
                        st.rerun()

    # ==============================
    # TAB ACTIONS
    # ==============================
    with tab_actions:
        st.markdown("<h4 style='color: #0A192F;'>🔄 Multi-Stage Loan Rollover</h4>", unsafe_allow_html=True)
        eligible_loans = loans_df[(~loans_df["status"].isin(["CLEARED"])) & (loans_df["balance"] > 0)]
    
        if eligible_loans.empty:
            st.success("All loans brought up to date! ✨")
        else:
            roll_map = {f"{row['borrower']} • {row['loan_id_label']} • Cycle {row['cycle_no']} • Bal {row['balance']:,.0f}": row["id"] for _, row in eligible_loans.iterrows()}
            roll_sel = st.selectbox("Select Loan to Roll Forward", list(roll_map.keys()))
            parent_id = roll_map[roll_sel]
    
            loan_to_roll = eligible_loans[eligible_loans["id"] == parent_id].iloc[0]
            new_interest_rate = st.number_input("New Monthly Interest (%)", value=3.0, step=0.5)
    
            if st.button("🔥 Execute Next Rollover", use_container_width=True):
                old_due = pd.to_datetime(loan_to_roll["end_date"], errors="coerce")
                if pd.isna(old_due):
                    old_due = datetime.now()
    
                new_start = old_due
                new_due = old_due + timedelta(days=30)
                current_status = str(loan_to_roll["status"]).strip().upper()
    
                if current_status == "PENDING":
                    loans_df.loc[loans_df["id"] == parent_id, "status"] = "BCF"
    
                save_data_saas("loans", loans_df)
                unpaid = float(loan_to_roll["balance"])
                new_interest = unpaid * (new_interest_rate / 100)
    
                new_row = {
                    "id": str(uuid.uuid4()),
                    "sn": "PENDING",
                    "loan_id_label": "PENDING",
                    "parent_loan_id": parent_id,
                    "borrower_id": loan_to_roll["borrower_id"],
                    "loan_type": loan_to_roll["loan_type"],
                    "principal": unpaid, 
                    "interest": new_interest,
                    "total_repayable": unpaid + new_interest,
                    "amount_paid": 0.0,
                    "balance": unpaid + new_interest,
                    "status": "PENDING",
                    "start_date": str(new_start.date()),
                    "end_date": str(new_due.date()),
                    "cycle_no": int(loan_to_roll["cycle_no"]) + 1,
                    "tenant_id": get_current_tenant()
                }
    
                if save_data("loans", pd.DataFrame([new_row])):
                    st.success("✅ Loan rolled forward.")
                    st.cache_data.clear()
                    st.rerun()

    # ==============================
    # TAB MANAGE
    # ==============================
    with tab_manage:
        if not loans_df.empty:
            edit_map = {f"{row['borrower']} • {row['loan_id_label']} • Cycle {row['cycle_no']}": row["id"] for _, row in loans_df.iterrows()}
            selected = st.selectbox("Select Loan to Edit", list(edit_map.keys()))
            target_id = edit_map[selected]
    
            loan_match = loans_df[loans_df["id"] == target_id]
            if loan_match.empty:
                st.error("Loan not found.")
                st.stop()
    
            loan_to_edit = loan_match.iloc[0]
    
            with st.form(f"edit_form_{target_id}"):
                col1, col2 = st.columns(2)
                with col1:
                    e_princ = st.number_input("Principal", value=float(loan_to_edit["principal"]))
                with col2:
                    current_interest = float(loan_to_edit["interest"]) if pd.notna(loan_to_edit["interest"]) else 0.0
                    e_interest = st.number_input("Interest Amount", value=current_interest)
    
                col3, col4 = st.columns(2)
                with col3:
                    current_start_date = pd.to_datetime(loan_to_edit["start_date"]).date() if pd.notna(loan_to_edit["start_date"]) else pd.Timestamp.now().date()
                    e_start_date = st.date_input("Start Date", value=current_start_date)
    
                with col4:
                    current_end_date = pd.to_datetime(loan_to_edit["end_date"]).date() if pd.notna(loan_to_edit["end_date"]) else pd.Timestamp.now().date()
                    e_end_date = st.date_input("End Date", value=current_end_date)
    
                status_options = ["ACTIVE", "PENDING", "CLEARED", "BCF", "CLOSED"]
                current_stat = str(loan_to_edit["status"]).upper()
                idx = status_options.index(current_stat) if current_stat in status_options else 0
    
                e_stat = st.selectbox("Status", status_options, index=idx)
    
                if st.form_submit_button("💾 Save Changes"):
                    # 🎯 FIX 3: Recompute total_repayable dynamically on submit to keep DB accurate
                    updated_total_repayable = float(e_princ + e_interest)
                    
                    supabase.table("loans").update(
                        {
                            "principal": e_princ,
                            "interest": e_interest,
                            "total_repayable": updated_total_repayable,
                            "start_date": e_start_date.isoformat(),
                            "end_date": e_end_date.isoformat(),
                            "status": e_stat,
                        }
                    ).eq("id", target_id).execute()
    
                    st.success("✅ Updated!")
                    st.cache_data.clear()
                    st.rerun()
    
            if st.button("🗑️ Delete Loan Permanently", use_container_width=True):
                supabase.table("loans").delete().eq("id", target_id).execute()
                st.warning("Loan Deleted.")
                st.cache_data.clear()
                st.rerun()

# ==============================
# 🧾 RECEIPT GENERATION & UTILS
# ==============================
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from io import BytesIO
from datetime import datetime
import pandas as pd
import streamlit as st

def generate_receipt_pdf(data):
    """
    Renders receipt documents directly to an in-memory buffer 
    to maximize transaction safety and performance across sessions.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    content = []
    
    content.append(Paragraph("<b>PAYMENT RECEIPT</b>", styles["Title"]))
    content.append(Spacer(1, 15))
    
    for k, v in data.items():
        content.append(Paragraph(f"<b>{k}:</b> {v}", styles["Normal"]))
        content.append(Spacer(1, 10))
        
    doc.build(content)
    buffer.seek(0)
    return buffer

# ✅ SINGLE SOURCE OF TRUTH (RPC)
def generate_receipt_no(supabase, tenant_id):
    try:
        res = supabase.rpc("get_next_receipt", {"p_tenant": tenant_id}).execute()
        if res.data:
            return res.data
        return f"RCPT-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    except Exception as e:
        return f"RCPT-{datetime.now().strftime('%Y%m%d%H%M%S')}"

# ==============================
# 💵 PAYMENTS MODULE (CYCLE-AWARE)
# ==============================
def show_payments(supabase):
    brand_color = st.session_state.get("theme_color", "#2B3F87")
    current_tenant = st.session_state.get('tenant_id')

    if not current_tenant:
        st.error("🔐 Session expired. Please log in.")
        st.stop()

    st.markdown(f"<h2 style='color: {brand_color}; margin-bottom: 20px;'>💵 Payments Management</h2>", unsafe_allow_html=True)

    try:
        loans_raw = get_cached_data("loans")
        payments_raw = get_cached_data("payments")
        borrowers_raw = get_cached_data("borrowers")
    except Exception as e:
        st.error(f"❌ Data load error: {e}")
        return

    # Structure dataframes safely
    raw_loans_df = pd.DataFrame(loans_raw) if loans_raw is not None else pd.DataFrame()
    raw_payments_df = pd.DataFrame(payments_raw) if payments_raw is not None else pd.DataFrame()
    raw_borrowers_df = pd.DataFrame(borrowers_raw) if borrowers_raw is not None else pd.DataFrame()

    # Normalize column schemas early
    for dataframe in [raw_loans_df, raw_payments_df, raw_borrowers_df]:
        if not dataframe.empty:
            dataframe.columns = dataframe.columns.str.lower().str.strip().str.replace(" ", "_")

    if raw_loans_df.empty:
        st.info("ℹ️ No active corporate loan records discoverable on the network.")
        return

    # 🛡️ TENANT FILTER BOUNDARY: Ensure strict company isolation
    if "tenant_id" in raw_loans_df.columns:
        loans_df = raw_loans_df[raw_loans_df["tenant_id"].astype(str) == str(current_tenant)].copy()
    else:
        loans_df = raw_loans_df.copy()

    if loans_df.empty:
        st.info("ℹ️ No active loan parameters logged under this profile.")
        return

    payments_df = raw_payments_df if "tenant_id" not in raw_payments_df.columns else raw_payments_df[raw_payments_df["tenant_id"].astype(str) == str(current_tenant)].copy()
    borrowers_df = raw_borrowers_df if "tenant_id" not in raw_borrowers_df.columns else raw_borrowers_df[raw_borrowers_df["tenant_id"].astype(str) == str(current_tenant)].copy()

    # Ensure clean string casting for relationship keys
    for df, col in [(borrowers_df, "id"), (loans_df, "borrower_id"), (loans_df, "id"), (payments_df, "loan_id")]:
        if col in df.columns:
            df[col] = df[col].astype(str)

    # Borrower profile transformations
    if not borrowers_df.empty and "name" in borrowers_df.columns:
        borrower_map = dict(zip(borrowers_df["id"], borrowers_df["name"]))
        loans_df["borrower"] = loans_df["borrower_id"].map(borrower_map).fillna("Unknown Profile")
    else:
        loans_df["borrower"] = "Unknown Profile"

    # Numeric Coercion
    numeric_cols = ["total_repayable", "amount_paid", "balance", "principal", "interest"]
    for col in numeric_cols:
        if col in loans_df.columns:
            loans_df[col] = pd.to_numeric(loans_df[col].fillna(0), errors="coerce")

    if not payments_df.empty and "amount" in payments_df.columns:
        payments_df["amount"] = pd.to_numeric(payments_df["amount"].fillna(0), errors="coerce")

    # Aggregate dynamically across matching ledgers
    if not payments_df.empty:
        payment_sums = payments_df.groupby("loan_id")["amount"].sum()
        loans_df["amount_paid"] = loans_df["id"].map(payment_sums).fillna(0)
    else:
        loans_df["amount_paid"] = 0

    loans_df["balance"] = loans_df["total_repayable"] - loans_df["amount_paid"]

    # Business Logic Utilities
    def cascade_payment(df_target, sn, changed_cycle_no):
        cycles = df_target[df_target["sn"] == sn].sort_values("cycle_no").copy()
        for idx, row in cycles.iterrows():
            if int(row["cycle_no"]) <= int(changed_cycle_no):
                continue
            pos = cycles.index.get_loc(idx)
            prev_idx = cycles.index[pos - 1]
            
            prev_balance = df_target.loc[prev_idx, "balance"]
            current_interest = row["interest"]
            
            df_target.loc[idx, "principal"] = prev_balance
            df_target.loc[idx, "total_repayable"] = prev_balance + current_interest
            df_target.loc[idx, "balance"] = df_target.loc[idx, "total_repayable"] - df_target.loc[idx, "amount_paid"]

    def get_active_loan(df_source, loan_row):
        current = loan_row
        visited = set()
        while True:
            if current["id"] in visited:
                break
            visited.add(current["id"])
            child = df_source[df_source["parent_loan_id"] == current["id"]]
            if child.empty:
                return current
            current = child.iloc[0]
        return current

    # ==============================
    # 📑 INTERFACE NAVIGATION TABS
    # ==============================
    tab1, tab2 = st.tabs(["➕ Record Payment Collection", "📜 Repayment History Logs"])

    with tab1:
        active_loans = loans_df.copy()
        active_loans["label"] = active_loans.apply(
            lambda r: f"{r['borrower']} | Ref: {r.get('loan_id_label', r['id'][:8])} | Owed: UGX {r['balance']:,.0f}", axis=1
        )

        selected_index = st.selectbox("🎯 Choose Target Active Account Line", active_loans.index, format_func=lambda i: active_loans.loc[i, "label"])
        loan = active_loans.loc[selected_index]
        active_loan = get_active_loan(loans_df, loan)
        loan_id = active_loan["id"]
        current_bal = active_loan["total_repayable"] - active_loan["amount_paid"]

        st.info(f"Targeting Account Pipeline: {active_loan['borrower']} (Ref ID: {loan_id[:8]})")
        st.metric("Verified Balance Position", f"UGX {current_bal:,.0f}")

        with st.form("payment_form"):
            amount = st.number_input("Transaction Amount Received (UGX)", min_value=0.0, step=10000.0) # Matched floating signatures
            method = st.selectbox("Collection Channel / Method", ["Cash", "Mobile Money", "Bank"])
            date = st.date_input("Processing Settlement Date", datetime.now())
            submit = st.form_submit_button("🚀 Post Repayment Entry", use_container_width=True)

        if submit:
            if amount <= 0:
                st.warning("⚠️ Transaction entries require a positive value contribution.")
            else:
                try:
                    receipt_no = generate_receipt_no(supabase, current_tenant)

                    # 1️⃣ Insert backend payment transaction
                    supabase.table("payments").insert({
                        "receipt_no": receipt_no,
                        "loan_id": loan_id,
                        "borrower": active_loan["borrower"],
                        "amount": float(amount),
                        "date": date.strftime("%Y-%m-%d"),
                        "method": method,
                        "tenant_id": current_tenant
                    }).execute()

                    # 2️⃣ Update state matrices
                    loans_df.loc[loans_df["id"] == loan_id, "amount_paid"] += amount
                    loans_df.loc[loans_df["id"] == loan_id, "balance"] = (
                        loans_df.loc[loans_df["id"] == loan_id, "total_repayable"] - loans_df.loc[loans_df["id"] == loan_id, "amount_paid"]
                    )

                    # 3️⃣ Cascade interest balance implications
                    cascade_payment(loans_df, active_loan["sn"], int(active_loan["cycle_no"]))
                    save_data_saas("loans", loans_df)

                    # 4️⃣ Generate memory-isolated receipts cleanly
                    pdf_buffer = generate_receipt_pdf({
                        "Receipt Number": receipt_no,
                        "Borrower Entity": active_loan["borrower"],
                        "Settlement Amount": f"UGX {amount:,.0f}",
                        "Payment Framework": method,
                        "Execution Date": date.strftime("%Y-%m-%d"),
                    })

                    st.download_button(
                        label="📥 Download Formal Receipt Blueprint", 
                        data=pdf_buffer, 
                        file_name=f"Receipt_{receipt_no}.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )

                    st.success(f"✅ Payment registered. Adjusted Owed Position: UGX {loans_df.loc[loans_df['id'] == loan_id, 'balance'].values[0]:,.0f}")
                    st.cache_data.clear()
                    st.rerun()

                except Exception as e:
                    st.error(f"❌ Post transaction sequence aborted: {e}")

    with tab2:
        if payments_df.empty:
            st.info("No localized accounting receipts currently matched inside this ledger profile.")
        else:
            # Create a clean, separate copy for layout mutations to prevent indexing side-effects
            display_df = payments_df.copy()
            display_df["amount_display"] = display_df["amount"].apply(lambda x: f"UGX {x:,.0f}")
            display_df["receipt_no"] = display_df["receipt_no"].fillna("Pending Log Ref")
            
            date_col = "date" if "date" in display_df.columns else "payment_date"
            display_df = display_df.sort_values(by=date_col, ascending=False)
            
            st.dataframe(
                display_df[[date_col, "borrower", "amount_display", "method", "receipt_no"]], 
                use_container_width=True, 
                hide_index=True
            )

            st.markdown("---")
            st.markdown("### ⚙️ Ledger Log Maintenance Modality")

            # Map selection using unique string IDs
            pay_map = {
                f"{row['receipt_no']} | {row['borrower']} | {row['amount_display']}": str(row['id'])
                for _, row in display_df.iterrows()
            }

            selected_pay_label = st.selectbox("Choose Targeted Transaction Allocation", list(pay_map.keys()))
            target_pay_id = pay_map[selected_pay_label]
            
            # SAFE FILTERING LAYER: Avoid layout index crashes
            matched_rows = display_df[display_df['id'].astype(str) == str(target_pay_id)]
            
            if matched_rows.empty:
                st.error("⚠️ Selected payment records could not be resolved from the current data frame.")
            else:
                # Safely extract target sequence without using rigid positional index assumptions
                target_pay = matched_rows.iloc[0]

                p_col1, p_col2 = st.columns(2)

                if p_col1.button("🗑️ Purge Payment From Records", use_container_width=True):
                    try:
                        # 1️⃣ Execute Supabase backend mutation drop
                        supabase.table("payments").delete().eq("id", target_pay_id).execute()
                        
                        loan_id = str(target_pay["loan_id"])
                        
                        # Guard against downstream lookups failing if loan object metadata structure is modified
                        matching_loans = loans_df[loans_df["id"].astype(str) == loan_id]
                        if not matching_loans.empty:
                            affected_loan = matching_loans.iloc[0]
                            
                            # 2️⃣ Readjust running balance calculations locally
                            loans_df.loc[loans_df["id"].astype(str) == loan_id, "amount_paid"] -= float(target_pay["amount"])
                            loans_df.loc[loans_df["id"].astype(str) == loan_id, "balance"] = (
                                loans_df.loc[loans_df["id"].astype(str) == loan_id, "total_repayable"] - 
                                loans_df.loc[loans_df["id"].astype(str) == loan_id, "amount_paid"]
                            )
                            
                            # 3️⃣ Cascade interest structural sequences down-chain
                            cascade_payment(loans_df, affected_loan["sn"], int(affected_loan["cycle_no"]))
                            save_data_saas("loans", loans_df)
                        
                        st.cache_data.clear()
                        st.warning(f"Payment reference key {target_pay['receipt_no']} dropped from backend successfully.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Execution aborted: {e}")

                if p_col2.button("📝 Edit Asset Metrics", use_container_width=True):
                    st.session_state["edit_pay_mode"] = True

                if st.session_state.get("edit_pay_mode"):
                    with st.form("edit_payment_form"):
                        st.info(f"Modifying Entry Core Attributes: {target_pay['receipt_no']}")
                        
                        # Float value types calibrated with explicit steps to bypass number input crash risks
                        new_amt = st.number_input("Revised Amount Allocation (UGX)", value=float(target_pay['amount']), step=5000.0)
                        
                        current_method = target_pay['method']
                        method_options = ["Cash", "Mobile Money", "Bank"]
                        method_idx = method_options.index(current_method) if current_method in method_options else 0
                        new_method = st.selectbox("Revised Collection Channel", method_options, index=method_idx)
                        
                        eb1, eb2 = st.columns(2)

                        if eb1.form_submit_button("💾 Save Adjustment Matrices", use_container_width=True):
                            try:
                                supabase.table("payments").update({
                                    "amount": new_amt,
                                    "method": new_method
                                }).eq("id", target_pay_id).execute()
                                
                                loan_id = str(target_pay["loan_id"])
                                matching_loans = loans_df[loans_df["id"].astype(str) == loan_id]
                                
                                if not matching_loans.empty:
                                    affected_loan = matching_loans.iloc[0]
                                    
                                    diff = new_amt - float(target_pay["amount"])
                                    loans_df.loc[loans_df["id"].astype(str) == loan_id, "amount_paid"] += diff
                                    loans_df.loc[loans_df["id"].astype(str) == loan_id, "balance"] = (
                                        loans_df.loc[loans_df["id"].astype(str) == loan_id, "total_repayable"] - 
                                        loans_df.loc[loans_df["id"].astype(str) == loan_id, "amount_paid"]
                                    )
                                    
                                    cascade_payment(loans_df, affected_loan["sn"], int(affected_loan["cycle_no"]))
                                    save_data_saas("loans", loans_df)
                                
                                st.session_state["edit_pay_mode"] = False
                                st.cache_data.clear()
                                st.success("Changes committed.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Update failed: {e}")

                        if eb2.form_submit_button("❌ Drop Corrections", use_container_width=True):
                            st.session_state["edit_pay_mode"] = False
                            st.rerun()

import streamlit as st
import pandas as pd
import uuid
from datetime import datetime
from io import BytesIO

def format_with_commas(df):
    if df.empty:
        return df
    df = df.copy()
    numeric_cols = df.select_dtypes(include=["number"]).columns
    for col in numeric_cols:
        df[col] = df[col].apply(lambda x: f"{int(x):,}" if pd.notnull(x) else "")
    return df

def delete_data_saas(table_name, filters):
    try:
        response = supabase.table(table_name).delete().match(filters).execute()
        return True
    except Exception as e:
        st.error(f"Database Error: {e}")
        return False

# =================================
# 🏢 Enterprise Payroll Engine (Compliant Excel Export)
# =================================
def export_styled_excel(df, company="ZOE CONSULTS SMC LTD"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"

    # Styles
    blue = PatternFill("solid", fgColor="4A90E2")
    white_font = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right")

    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells("A1:W1")
    ws["A1"] = f"{datetime.now().strftime('%B %Y').upper()} PAYROLL ({company})"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    # Header Row 1
    ws.append([
        "S/NO","Employee Name","TIN","Designation","Mob No",
        "Account No","NSSF No",
        "Salary","Basic",
        "NO PAY","",
        "LST",
        "Gross Salary",
        "Deductions",
        "If yes calculate","","",
        "Other",
        "Total Deductions",
        "Nett Pay",
        "Total Tax",
        "10% NSSF",
        "15% NSSF"
    ])
    ws.merge_cells("J2:K2")
    ws.merge_cells("O2:Q2")

    # Header Row 2
    ws.append([
        "No","","","","","","",
        "ARREARS","Salary",
        "DAYS","Absenteeism",
        "Deductions",
        "",
        "P.A.Y.E","N.S.S.F","S.DRS/ADV",
        "Deduction",
        "", "", "", "", ""
    ])

    for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=23):
        for cell in row:
            cell.fill = blue
            cell.font = white_font
            cell.alignment = center
            cell.border = thin

    # Data - FIXED COLUMN MAPPING CORRECTION
    for idx, r in df.reset_index(drop=True).iterrows():
        ws.append([
            idx + 1,
            r["employee"],
            r["tin"],
            r["designation"],
            r["mob_no"],
            r["account_no"],
            r["nssf_no"],
            r["arrears"],
            r["basic_salary"],
            0,
            r["absent_deduction"],
            r["lst"],
            r["gross_salary"],
            r["paye"],           # P.A.Y.E
            r["nssf_5"],         # N.S.S.F (5% Employee contribution)
            r["advance_drs"],    # S.DRS/ADV
            r["other_deductions"],
            r["paye"] + r["nssf_5"] + r["advance_drs"] + r["other_deductions"] + r["lst"], # Total Deductions
            r["net_pay"],        # Nett Pay
            r["paye"],           # Total Tax
            r["nssf_10"],        # 10% Employer NSSF
            r["nssf_15"]         # 15% Total NSSF
        ])

    # Number formatting
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=8, max_col=23):
        for cell in row:
            cell.number_format = '#,##0'
            cell.alignment = right
            cell.border = thin

    # Totals
    total_row = ws.max_row + 1
    ws[f"A{total_row}"] = "TOTAL"

    for col in range(8, 24):
        letter = get_column_letter(col)
        ws[f"{letter}{total_row}"] = f"=SUM({letter}4:{letter}{total_row-1})"

    for col in range(1, 24):
        cell = ws[f"{get_column_letter(col)}{total_row}"]
        cell.font = bold
        cell.fill = blue
        cell.border = thin

    for col in range(8, 24):
        cell = ws[f"{get_column_letter(col)}{total_row}"]
        cell.number_format = '#,##0'
        cell.alignment = right

    widths = [6,22,15,25,15,20,20] + [12]*16
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ---------------------------------
# Legal Ugandan Calculation Engine
# ---------------------------------
def compute_payroll(basic, arrears, absent, advance, other, apply_lst=True):
    # Gross salary = Earnings minus structural absenteeism penalties
    gross = round(float(basic) + float(arrears) - float(absent))

    # NSSF Standard Uganda (Based cleanly on Gross)
    nssf_5 = round(gross * 0.05)
    nssf_10 = round(gross * 0.10)
    nssf_15 = nssf_5 + nssf_10

    # COMPLIANCE FIX: In Uganda, PAYE base is Gross Salary (Employee 5% NSSF is NOT deductible)
    taxable_income = gross 

    # COMPLIANCE FIX: Legal URA PAYE Brackets including the >10M Tier (Additional 10%)
    paye = 0
    if taxable_income <= 235000:
        paye = 0
    elif taxable_income <= 335000:
        paye = (taxable_income - 235000) * 0.10
    elif taxable_income <= 410000:
        paye = 10000 + (taxable_income - 335000) * 0.20
    elif taxable_income <= 10000000:
        paye = 25000 + (taxable_income - 410000) * 0.30
    else:
        # Tier 4 Luxury Bracket: Standard Tier 3 baseline + 30% + 10% additional surtax
        standard_tier3 = 25000 + (taxable_income - 410000) * 0.30
        additional_surtax = (taxable_income - 10000000) * 0.10
        paye = standard_tier3 + additional_surtax

    paye = round(paye)

    # Local Service Tax (LST Rules Uganda)
    lst = 0
    if apply_lst and (gross > 100000): # Statutory lower bound threshold
        if gross <= 300000:
            lst = round(5000 / 12)
        elif gross <= 400000:
            lst = round(10000 / 12)
        elif gross <= 600000:
            lst = round(20000 / 12)
        elif gross <= 1000000:
            lst = round(30000 / 12)
        else:
            lst = round(100000 / 12)

    total_deductions = paye + nssf_5 + advance + other + lst
    net = gross - total_deductions

    return {
        "gross": gross,
        "lst": lst,
        "n5": nssf_5,
        "n10": nssf_10,
        "n15": nssf_15,
        "paye": paye,
        "net": round(net)
    }

def format_payroll_display(df):
    if df.empty:
        return df

    df = df.copy()
    df["NO"] = range(1, len(df) + 1)
    df["Salary"] = df["gross_salary"]
    df["Basic"] = df["basic_salary"]
    df["NO PAY DAYS"] = 0
    df["Absenteeism"] = df["absent_deduction"]
    df["Gross Salary"] = df["gross_salary"]
    df["Deductions"] = df["paye"]
    df["P.A.Y.E"] = df["paye"]
    df["N.S.S.F"] = df["nssf_5"]
    df["S.DRS/ADV"] = df["advance_drs"]
    df["Other"] = df["other_deductions"]
    df["Total Deductions"] = df["paye"] + df["nssf_5"] + df["advance_drs"] + df["other_deductions"] + df["lst"]
    df["Nett Pay"] = df["net_pay"]
    df["Total Tax on Salary"] = df["paye"]
    df["10% NSSF"] = df["nssf_10"]
    df["15% NSSF"] = df["nssf_15"]

    return df[[
        "NO", "employee", "tin", "designation", "mob_no", "account_no", "nssf_no",
        "Salary", "Basic", "NO PAY DAYS", "Absenteeism", "Gross Salary", "Deductions",
        "P.A.Y.E", "N.S.S.F", "S.DRS/ADV", "Other", "Total Deductions", "Nett Pay",
        "Total Tax on Salary", "10% NSSF", "15% NSSF"
    ]]

def show_payroll():
    tenant = st.session_state.get("tenant_id")
    role = st.session_state.get("role")

    if not tenant or role != "Admin":
        st.error("🔒 Restricted: Access requires Administrator elevation.")
        return

    st.markdown("<h2 style='color:#4A90E2;'>🧾 Payroll</h2>", unsafe_allow_html=True)

    payroll_df = get_cached_data("payroll")
    if payroll_df is not None and not payroll_df.empty:
        payroll_df.columns = payroll_df.columns.astype(str).str.strip().str.replace(" ", "_")
        payroll_df = payroll_df[payroll_df["tenant_id"].astype(str) == str(tenant)]
    else:
        payroll_df = pd.DataFrame()

    employee_list = []
    if not payroll_df.empty and "employee" in payroll_df.columns:
        employee_list = sorted(payroll_df["employee"].dropna().astype(str).unique())

    tab_process, tab_history = st.tabs(["💳 Process Payroll", "📜 Payroll History"])

    with tab_process:
        # UI Safety Container to encapsulate errors without cutting off app framework execution
        validation_failed = False
        
        with st.form("payroll_form", clear_on_submit=True):
            st.subheader("👤 Employee Info")
            selected_emp = st.selectbox("Select Employee", employee_list) if employee_list else None
            new_emp = st.text_input("Or Enter New Employee")
            employee_name = new_emp if new_emp else selected_emp

            c1, c2, c3 = st.columns(3)
            f_tin = c1.text_input("TIN")
            f_desig = c2.text_input("Designation")
            f_mob = c3.text_input("Mobile")

            c4, c5 = st.columns(2)
            f_acc = c4.text_input("Account No")
            f_nssf = c5.text_input("NSSF No")

            st.subheader("💰 Earnings & Deductions")
            f_apply_lst = st.checkbox("Deduct Local Service Tax (LST)?", value=True)

            c6, c7, c8 = st.columns(3)
            f_arrears = c6.number_input("Arrears", min_value=0.0)
            f_basic = c7.number_input("Basic Salary", min_value=0.0)
            f_absent = c8.number_input("Absent Deduction", min_value=0.0)

            c9, c10 = st.columns(2)
            f_adv = c9.number_input("Advance", min_value=0.0)
            f_other = c10.number_input("Other Deductions", min_value=0.0)

            if st.form_submit_button("💳 Save Payroll"):
                if not employee_name or f_basic <= 0:
                    st.error("⚠️ Entry Blocked: Valid employee selection and non-zero baseline basic salary required.")
                    validation_failed = True
                
                month_str = datetime.now().strftime("%Y-%m")
                if not payroll_df.empty and not validation_failed:
                    duplicate = payroll_df[(payroll_df["employee"] == employee_name) & (payroll_df["month"] == month_str)]
                    if not duplicate.empty:
                        st.warning("⚠️ Warning: Statement submission skipped. Payroll run already registered for this month.")
                        validation_failed = True

                if not validation_failed:
                    calc = compute_payroll(f_basic, f_arrears, f_absent, f_adv, f_other, apply_lst=f_apply_lst)

                    new_row = pd.DataFrame([{
                        "payroll_id": str(uuid.uuid4()),
                        "employee": employee_name,
                        "tin": f_tin,
                        "designation": f_desig,
                        "mob_no": f_mob,
                        "account_no": f_acc,
                        "nssf_no": f_nssf,
                        "arrears": f_arrears,
                        "basic_salary": f_basic,
                        "absent_deduction": f_absent,
                        "gross_salary": calc["gross"],
                        "lst": calc["lst"],
                        "paye": calc["paye"],
                        "nssf_5": calc["n5"],
                        "nssf_10": calc["n10"],
                        "nssf_15": calc["n15"],
                        "advance_drs": f_adv,
                        "other_deductions": f_other,
                        "net_pay": calc["net"],
                        "date": datetime.utcnow().isoformat(),
                        "month": month_str,
                        "tenant_id": str(tenant)
                    }])

                    if save_data_saas("payroll", new_row):
                        st.cache_data.clear()
                        st.success(f"✅ Saved completely for {employee_name}")
                        st.rerun()

    with tab_history:
        if payroll_df.empty:
            st.info("No current payroll entries found.")
            return

        st.markdown("### 📊 Payroll Sheet")
        display_df = format_payroll_display(payroll_df)
        
        try:
            formatted_df = format_with_commas(display_df)
        except NameError:
            formatted_df = display_df 

        st.dataframe(formatted_df, use_container_width=True)

        st.markdown("---")
        st.subheader("🛠️ Manage Records")
        
        record_options = payroll_df.apply(lambda x: f"{x['employee']} ({x['month']}) | ID: {x['payroll_id'][:8]}", axis=1).tolist()
        selected_record_str = st.selectbox("Select a record to Edit or Delete", options=record_options)

        if selected_record_str:
            sel_id = selected_record_str.split("| ID: ")[1].strip()
            full_record = payroll_df[payroll_df['payroll_id'].str.contains(sel_id)].iloc[0]

            col_edit, col_del = st.columns(2)
            with col_edit:
                if st.button("📝 Edit Selected Record"):
                    st.warning("To override context: Re-submit entry on processing engine profile utilizing duplicate target month identifier parameters.")
            
            with col_del:
                if st.button("🗑️ Delete Record", type="primary"):
                    if delete_data_saas("payroll", {"payroll_id": full_record['payroll_id']}):
                        st.cache_data.clear()
                        st.success(f"Deleted payroll allocation for {full_record['employee']}")
                        st.rerun()

        st.markdown("---")
        c1, c2 = st.columns(2)
        with c1:
            csv = payroll_df.to_csv(index=False).encode("utf-8")
            st.download_button(
                "📄 Download CSV",
                data=csv,
                file_name=f"Payroll_{datetime.now().strftime('%Y%m%d')}.csv",
                use_container_width=True
            )
        with c2:
            excel_file = export_styled_excel(payroll_df)
            st.download_button(
                "📥 Download Styled Excel",
                data=excel_file,
                file_name=f"Payroll_Styled_{datetime.now().strftime('%B_%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
# ==========================================================
# 🚀 BALLISTIC FINTECH REPORTS ENGINE (EXPENSE-OPTIMIZED)
# ==========================================================
import streamlit as st
import pandas as pd
import plotly.express as px 
from datetime import datetime

def show_reports():
    """
    Advanced financial reporting with multi-tenant isolation 
    and investor-grade intelligence metrics.
    """

    # ==============================
    # 🎨 HEADER (EXECUTIVE UI)
    # ==============================
    st.markdown("""
    <div style='background: linear-gradient(90deg,#1E3A8A,#2B3F87); padding:20px; border-radius:15px; margin-bottom:25px;'>
        <h2 style='margin:0; color:white; font-size:24px;'>📊 Financial Intelligence Dashboard</h2>
        <p style='margin:0; color:#DBEAFE; font-size:13px;'>Real-time P&L, Balance Sheet, and Portfolio Yield Analysis</p>
    </div>
    """, unsafe_allow_html=True)

    tenant = st.session_state.get("tenant_id")
    if not tenant:
        st.error("Session Expired.")
        return

    # ==============================
    # 🛡️ DATA FETCH & TENANT SAFETY
    # ==============================
    def safe_tenant_filter(df_name):
        try:
            df = get_cached_data(df_name)
            if df is None or df.empty:
                return pd.DataFrame()
            if "tenant_id" in df.columns:
                return df[df["tenant_id"].astype(str) == str(tenant)].copy()
            return df
        except Exception:
            return pd.DataFrame()

    loans = safe_tenant_filter("loans")
    payments = safe_tenant_filter("payments")
    expenses = safe_tenant_filter("expenses")
    borrowers = safe_tenant_filter("borrowers")

    if loans.empty:
        st.info("💡 No loan data found. Financial reports will populate once loans are issued.")
        return

    # ==============================
    # 🧰 REUSABLE HELPERS
    # ==============================
    def col_sum(df, col):
        if df.empty or col not in df.columns: return 0.0
        return pd.to_numeric(df[col], errors="coerce").fillna(0).sum()

    def attach_borrower_names(loans_df, borrowers_df):
        if borrowers_df.empty or "id" not in borrowers_df.columns or "name" not in borrowers_df.columns:
            loans_df["borrower"] = loans_df.get("borrower", "Unknown")
            return loans_df
        borrowers_df["id"] = borrowers_df["id"].astype(str)
        bor_map = dict(zip(borrowers_df["id"], borrowers_df["name"]))
        mapped = loans_df["borrower_id"].map(bor_map)
        loans_df["borrower"] = mapped.fillna(loans_df.get("borrower")).fillna("Unknown")
        return loans_df

    loans = attach_borrower_names(loans, borrowers)

    # Ensure numeric safety
    for col in ["principal", "interest", "cycle_no", "balance"]:
        if col in loans.columns:
            loans[col] = pd.to_numeric(loans[col], errors="coerce").fillna(0)
    
    loans["status"] = loans["status"].astype(str).str.upper()
    
    # ==============================
    # 🧠 ACTIVE CAPITAL (Cycle 1 ONLY)
    # ==============================
    active_capital_loans = loans[
        (loans["cycle_no"] == 1) &
        (loans["status"].isin(["ACTIVE", "PENDING"]))
    ]
    total_capital_out = active_capital_loans["principal"].sum()
    
    # ==============================
    # 💰 INT. REVENUE (ONLY CLEARED LOANS)
    # ==============================
    cleared_loans = loans[loans["status"] == "CLEARED"]
    projected_interest = cleared_loans["interest"].sum()
    
    # ==============================
    # 💸 GLOBAL OPEX BREAKDOWN (FIXED ONE-SOURCE)
    # ==============================
    actual_collected = col_sum(payments, "amount")
    
    # All operational tracking (Salaries, NSSF, PAYE, Rent, etc.) are already bundled here
    total_direct_expenses = col_sum(expenses, "amount")
    
    # Single source operational expense tracking
    global_opex = total_direct_expenses
    cash_profit = actual_collected - global_opex
    
    # ==============================
    # 💎 KPI TILES
    # ==============================
    def render_kpi(title, value, color, icon="💰"):
        st.markdown(f"""
            <div style="padding:16px; border-radius:12px; background:white; border:1px solid #E5E7EB; box-shadow:0 2px 4px rgba(0,0,0,0.02); margin-bottom:10px;">
                <p style="font-size:11px; color:#6B7280; margin:0; font-weight:600;">{icon} {title}</p>
                <h3 style="margin:0; color:{color}; font-size:20px;">UGX {value:,.0f}</h3>
            </div>
        """, unsafe_allow_html=True)

    k1, k2, k3, k4 = st.columns(4)
    with k1: render_kpi("ACTIVE CAPITAL", total_capital_out, "#1E3A8A")
    with k2: render_kpi("INT. REVENUE", projected_interest, "#059669")
    with k3: render_kpi("COLLECTIONS", actual_collected, "#7C3AED")
    with k4: render_kpi("NET CASHFLOW", cash_profit, "#059669" if cash_profit >= 0 else "#DC2626")

    # ==============================
    # 📈 TREND ANALYSIS
    # ==============================
    st.markdown("### 📈 Monthly Profit & Loss Trend", unsafe_allow_html=True)

    payments["date"] = pd.to_datetime(payments.get("date"), errors="coerce")
    expenses["date"] = pd.to_datetime(expenses.get("date"), errors="coerce")

    if not payments.empty:
        inc_m = payments.set_index("date").resample("ME")["amount"].sum()
        inc_m.name = "Income"
    else:
        inc_m = pd.Series(dtype=float, name="Income")

    if not expenses.empty:
        exp_m = expenses.set_index("date").resample("ME")["amount"].sum()
        exp_m.name = "Expenses"
    else:
        exp_m = pd.Series(dtype=float, name="Expenses")
        
    pl_combined = pd.concat([inc_m, exp_m], axis=1).fillna(0)
    pl_combined["Net"] = pl_combined["Income"] - pl_combined["Expenses"]

    if not pl_combined.empty and pl_combined.index.notnull().any():
        fig_trend = px.area(
            pl_combined,
            color_discrete_map={"Income": "#059669", "Expenses": "#EF4444", "Net": "#1E3A8A"},
            line_shape="spline",
            labels={"index":"Month", "value":"UGX"}
        )
        fig_trend.update_layout(
            hovermode="x unified",
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
        )
        st.plotly_chart(fig_trend, use_container_width=True)
    else:
        st.info("📉 No trend data available yet.")

    # ==============================
    # 🧠 INVESTOR INTELLIGENCE METRICS
    # ==============================
    overdue_loans = loans[loans["status"].str.contains("OVERDUE", na=False)]
    par_value = col_sum(overdue_loans, "balance")
    par_ratio = (par_value / total_capital_out * 100) if total_capital_out > 0 else 0
    yield_pct = (projected_interest / total_capital_out * 100) if total_capital_out > 0 else 0
    coll_eff = (actual_collected / (total_capital_out + projected_interest) * 100) if (total_capital_out + projected_interest) > 0 else 0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Portfolio Yield", f"{yield_pct:.1f}%")
    m2.metric("Collection Eff.", f"{coll_eff:.1f}%")
    m3.metric("PAR Ratio", f"{par_ratio:.1f}%", delta=f"{par_ratio:.1f}%", delta_color="inverse")
    m4.metric("OpEx Ratio", f"{(global_opex/actual_collected*100 if actual_collected > 0 else 0):.1f}%")

    # ==============================
    # 🧾 FINANCIAL STATEMENTS (CYCLE-AWARE)
    # ==============================
    def fiscal_year(dt):
        if pd.isna(dt): return "Unknown"
        return f"{dt.year}-{dt.year+1}" if dt.month >= 7 else f"{dt.year-1}-{dt.year}"
    
    loans["start_date"] = pd.to_datetime(loans.get("start_date"), errors="coerce")
    payments["date"] = pd.to_datetime(payments.get("date"), errors="coerce")
    expenses["date"] = pd.to_datetime(expenses.get("date"), errors="coerce")
    
    loans["fiscal_year"] = loans["start_date"].apply(fiscal_year)
    payments["fiscal_year"] = payments["date"].apply(fiscal_year)
    expenses["fiscal_year"] = expenses["date"].apply(fiscal_year)
    
    fiscal_years = sorted([fy for fy in loans["fiscal_year"].dropna().unique() if fy != "Unknown"])
    
    if not fiscal_years:
        st.info("⏳ Waiting for structured historical records to compile statement tabs.")
        return

    selected_fy = st.selectbox("Select Financial Year", fiscal_years)
    
    fy_loans = loans[loans["fiscal_year"] == selected_fy]
    fy_expenses = expenses[expenses["fiscal_year"] == selected_fy]
    fy_payments = payments[payments["fiscal_year"] == selected_fy]
    
    # ------------------------------
    # 💰 INCOME STATEMENT (P&L)
    # ------------------------------
    s1, s2 = st.columns(2)
    with s1:
        st.subheader(f"💰 Profit & Loss — FY {selected_fy}")
        
        fy_active_capital = fy_loans[
            (fy_loans["cycle_no"] == 1) & 
            (fy_loans["status"].isin(["ACTIVE", "PENDING"]))
        ]["principal"].sum()
        
        int_revenue = fy_loans[fy_loans["status"] == "CLEARED"]["interest"].sum()
        
        # Pulling unified calculations from expenses (Salaries + Tax already covered inside)
        fy_direct_exp = col_sum(fy_expenses, "amount")
        
        fy_total_opex = fy_direct_exp
        net_profit = int_revenue - fy_total_opex
        
        st.dataframe(pd.DataFrame({
            "Description": [
                "Active Capital (Cycle 1 ACTIVE/PENDING)",
                "Interest Revenue (CLEARED Loans)",
                "Total Operating Expenses (OPEX)",
                "Net Profit"
            ],
            "amount (UGX)": [
                f"{fy_active_capital:,.0f}",
                f"{int_revenue:,.0f}",
                f"{fy_total_opex:,.0f}",
                f"{net_profit:,.0f}"
            ]
        }), use_container_width=True)
    
    # ------------------------------
    # 🏦 BALANCE SHEET SNAPSHOT
    # ------------------------------
    with s2:
        st.subheader(f"🏦 Balance Sheet — FY {selected_fy}")
        
        # Outstanding Loan portfolio current asset valuation
        loan_book_value = fy_loans["balance"].sum()
        
        # Cash on Hand = Inflow collections minus outflow expenses
        cash_position = col_sum(fy_payments, "amount") - fy_total_opex
        
        # Total Asset valuation metric
        total_assets = loan_book_value + cash_position
        
        st.dataframe(pd.DataFrame({
            "Description": [
                "Loan Book Portfolio Value (Outstanding Principal + Accrued)",
                "Cash & Cash Equivalents",
                "Total Assets"
            ],
            "amount (UGX)": [
                f"{loan_book_value:,.0f}",
                f"{cash_position:,.0f}",
                f"{total_assets:,.0f}"
            ]
        }), use_container_width=True)
    
    # ==============================
    # 📤 EXPORT
    # ==============================
    with st.expander(f"📥 Export Executive Report — FY {selected_fy}"):
        export_rows = [{
            "Fiscal Year": selected_fy,
            "Active Capital": fy_active_capital,
            "Interest Revenue": int_revenue,
            "Total OPEX": fy_total_opex,
            "Net Profit": net_profit,
            "Cash Position": cash_position,
            "Loan Book Value": loan_book_value,
            "Total Assets": total_assets
        }]
        
        export_df = pd.DataFrame(export_rows)
        st.dataframe(export_df)
        
        csv = export_df.to_csv(index=False).encode("utf-8")
        st.download_button(
            label="⬇️ Download Full Executive Report (CSV)",
            data=csv,
            file_name=f"FinReport_{selected_fy}_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            use_container_width=True
        )

            
# ==========================================================
# 🚨 OVERDUE TRACKER & ACTIVITY CALENDAR ENGINE (WITH NAMES)
# ==========================================================
import streamlit as st
import pandas as pd
from datetime import datetime

try:
    from streamlit_calendar import calendar as st_calendar
except ImportError:
    def st_calendar(*args, **kwargs):
        st.warning("⚠️ 'streamlit-calendar' library missing. Please install it.")

def show_overdue_tracker():
    """
    Tracks overdue loans with AI-style risk scoring, tenant isolation,
    and clear borrower naming context.
    """
    brand_color = st.session_state.get("theme_color", "#2B3F87")
    current_tenant = st.session_state.get('tenant_id')

    # ==============================
    # 🎨 UI SYSTEM
    # ==============================
    st.markdown("""
    <style>
    .glass-card {
        backdrop-filter: blur(10px);
        background: linear-gradient(145deg, rgba(255,255,255,0.9), rgba(245,247,255,0.8));
        border-radius: 16px;
        padding: 20px;
        border: 1px solid rgba(239, 68, 68, 0.1);
        box-shadow: 0 4px 15px rgba(0,0,0,0.05);
    }
    .metric-title { font-size: 11px; color: #6b7280; font-weight: 600; text-transform: uppercase; }
    .metric-value { font-size: 24px; font-weight: 700; margin-top: 4px; }
    .ai-badge {
        background: #F0F4FF;
        color: #2B3F87;
        padding: 2px 8px;
        border-radius: 8px;
        font-size: 10px;
        font-weight: bold;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown(f"<h2 style='color:{brand_color};'>🚨 AI Overdue Intelligence</h2>", unsafe_allow_html=True)

    # ==============================
    # 📥 FETCH & UNIFY DATA
    # ==============================
    loans_df = get_cached_data("loans")
    borrowers_df = get_cached_data("borrowers")
    
    if loans_df is None or loans_df.empty:
        st.info("📅 No loan records found in the system.")
        return

    # 🛡️ Tenant Isolation
    loans_df = loans_df[loans_df["tenant_id"].astype(str) == str(current_tenant)].copy()
    
    if "due_date" not in loans_df.columns and "end_date" in loans_df.columns:
        loans_df["due_date"] = loans_df["end_date"]

    required_cols = ["id", "amount", "due_date", "borrower_id", "status"]
    for col in required_cols:
        if col not in loans_df.columns: loans_df[col] = None

    loans_df["amount"] = pd.to_numeric(loans_df["amount"], errors="coerce").fillna(0)
    loans_df["due_date"] = pd.to_datetime(loans_df["due_date"], errors="coerce")
    
    # 👤 MAPPED NAMES INTELLIGENCE
    if borrowers_df is not None and not borrowers_df.empty:
        borrowers_df['id'] = borrowers_df['id'].astype(str)
        loans_df['borrower_id'] = loans_df['borrower_id'].astype(str)
        bor_map = dict(zip(borrowers_df['id'], borrowers_df['name']))
        loans_df['borrower'] = loans_df['borrower_id'].map(bor_map).fillna("Unknown Borrower")
    else:
        loans_df['borrower'] = "Unknown Borrower"

    # Filter out finalized accounts
    active_df = loans_df[~loans_df["status"].astype(str).str.upper().isin(["PAID", "CLOSED", "CLEARED"])].copy()

    if active_df.empty:
        st.success("✅ Great job! All loans are currently up to date or cleared.")
        return

    # ==============================
    # 🧠 AI RISK SCORING ENGINE
    # ==============================
    today = pd.Timestamp.today().normalize()
    active_df["days_overdue"] = (today - active_df["due_date"]).dt.days
    overdue_df = active_df[active_df["days_overdue"] > 0].copy()

    if overdue_df.empty:
        st.success("🎉 No overdue payments detected today.")
        return

    def compute_risk_score(row):
        score = 0
        score += min(row["days_overdue"] * 1.5, 50) 
        if row["amount"] > 1000000: score += 30      
        elif row["amount"] > 500000: score += 20
        elif row["amount"] > 100000: score += 10
        if row["days_overdue"] > 30: score += 20    
        return min(score, 100)

    overdue_df["risk_score"] = overdue_df.apply(compute_risk_score, axis=1)

    def classify_risk(score):
        if score >= 75: return "🔴 High Risk"
        elif score >= 40: return "🟠 Watch"
        return "🟢 Stable"

    overdue_df["risk_level"] = overdue_df["risk_score"].apply(classify_risk)

    # ==============================
    # 💎 KPI DASHBOARD
    # ==============================
    total_count = len(overdue_df)
    critical_count = len(overdue_df[overdue_df["risk_score"] >= 75])
    total_exposure = overdue_df["amount"].sum()

    c1, c2, c3 = st.columns(3)
    c1.markdown(f"""<div class="glass-card"><div class="metric-title">Overdue Cases</div><div class="metric-value">{total_count}</div></div>""", unsafe_allow_html=True)
    c2.markdown(f"""<div class="glass-card"><div class="metric-title">Critical Attention</div><div class="metric-value" style="color:#EF4444;">{critical_count}</div></div>""", unsafe_allow_html=True)
    c3.markdown(f"""<div class="glass-card"><div class="metric-title">Total Exposure</div><div class="metric-value">UGX {total_exposure:,.0f}</div></div>""", unsafe_allow_html=True)

    # ==============================
    # 🔍 FILTER & SEARCH
    # ==============================
    st.markdown("<br>", unsafe_allow_html=True)
    f1, f2 = st.columns([1, 2])
    risk_filter = f1.selectbox("Filter Risk", ["All Levels", "🔴 High Risk", "🟠 Watch", "🟢 Stable"])
    search = f2.text_input("🔍 Search borrower or Loan ID", key="overdue_search_input")

    display_df = overdue_df.copy()
    if risk_filter != "All Levels":
        display_df = display_df[display_df["risk_level"] == risk_filter]
    if search:
        display_df = display_df[display_df.astype(str).apply(lambda row: row.str.lower().str.contains(search.lower()).any(), axis=1)]

    # ==============================
    # 🎨 AI RANKED TABLE (NOW WITH NAMES)
    # ==============================
    st.markdown("### 🔥 Collection Priority List")
    
    if not display_df.empty:
        st.dataframe(
            display_df.sort_values("risk_score", ascending=False)[
                ["borrower", "days_overdue", "amount", "risk_level", "risk_score", "id"]
            ].rename(columns={
                "borrower": "Borrower Name",  # Added Context
                "days_overdue": "Days Late",
                "amount": "Balance (UGX)",
                "risk_level": "Risk Level",
                "risk_score": "Score/100",
                "id": "Loan ID"
            }),
            use_container_width=True,
            hide_index=True
        )
    else:
        st.info("No records match the current filters.")

    # ==============================
    # 🧠 SMART INSIGHT PANEL (NOW WITH NAMES)
    # ==============================
    st.markdown("<br>", unsafe_allow_html=True)
    try:
        worst_case = overdue_df.loc[overdue_df["risk_score"].idxmax()]
        st.markdown(f"""
        <div class="glass-card" style="border-left: 5px solid #EF4444;">
            <span class="ai-badge">AI ANALYTICS</span><br><br>
            <b>Priority Alert:</b> Account <b>{worst_case['borrower']}</b> (Loan #{worst_case['id']}) requires immediate field intervention.<br>
            It is <b>{int(worst_case['days_overdue'])} days</b> overdue with a risk score of <b>{worst_case['risk_score']:.0f}/100</b>.<br>
            <p style="font-size:12px; color:#666; margin-top:10px;"><i>Strategy: Debt has entered the critical recovery phase. Check collateral status immediately.</i></p>
        </div>
        """, unsafe_allow_html=True)
    except Exception:
        pass

    # ==============================
    # ⚙️ QUICK ACTIONS (NOW WITH NAMES)
    # ==============================
    with st.expander("⚙️ Recovery Actions"):
        # Format the selection row to instantly communicate who the human borrower is
        action_options = overdue_df.apply(lambda x: f"Name: {x['borrower']} | ID: {x['id']} | Late: {x['days_overdue']} days", axis=1).tolist()
        
        if action_options:
            target_loan = st.selectbox("Select Loan to Action", action_options)
            
            # Extract out the ID correctly by capturing everything between "ID: " and the following break pipeline
            sel_id = target_loan.split(" | ID: ")[1].split(" | ")[0]
            sel_name = target_loan.split("Name: ")[1].split(" | ")[0]
            
            act1, act2 = st.columns(2)
            if act1.button("📞 Log Contact Made", use_container_width=True):
                st.toast(f"Contact log updated for {sel_name} (Loan #{sel_id})")
                
            if act2.button("✅ Mark Fully Recovered", use_container_width=True):
                update_data = pd.DataFrame([{"id": sel_id, "status": "Paid", "tenant_id": str(current_tenant)}])
                if save_data_saas("loans", update_data):
                    st.success(f"Loan #{sel_id} for {sel_name} moved to Paid status.")
                    st.cache_data.clear()
                    st.rerun()
        else:
            st.info("No overdue loans to process.")


import streamlit as st
import pandas as pd
from datetime import datetime

def show_calendar():
    """
    Displays the modern enterprise operation activity center, workload forecasts,
    and precise execution deadlines.
    """
    brand_color = st.session_state.get("theme_color", "#2B3F87")
    current_tenant = st.session_state.get('tenant_id')

    if not current_tenant:
        st.error("🔐 Session expired. Please log in.")
        st.stop()

    # ==========================================
    # UI STYLING ENGINE (ENTERPRISE GRADE)
    # ==========================================
    st.markdown(f"""
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap');
            
            /* Root adjustments */
            .main-title {{
                font-family: 'Plus Jakarta Sans', sans-serif;
                font-weight: 700;
                color: #111827;
                font-size: 28px;
                letter-spacing: -0.02em;
                margin-bottom: 4px;
            }}
            .subtitle {{
                font-family: 'Plus Jakarta Sans', sans-serif;
                color: #4B5563;
                font-size: 14px;
                margin-bottom: 24px;
            }}
            
            /* High-end Dashboard Cards */
            .kpi-container {{
                background: #FFFFFF;
                border: 1px solid #E5E7EB;
                border-radius: 16px;
                padding: 20px;
                box-shadow: 0px 1px 3px rgba(16, 24, 40, 0.1), 0px 1px 2px rgba(16, 24, 40, 0.06);
                transition: transform 0.2s ease, box-shadow 0.2s ease;
            }}
            .kpi-container:hover {{
                box-shadow: 0px 4px 6px -1px rgba(16, 24, 40, 0.1), 0px 2px 4px -1px rgba(16, 24, 40, 0.06);
            }}
            .kpi-label {{
                font-family: 'Plus Jakarta Sans', sans-serif;
                font-size: 12px;
                font-weight: 600;
                color: #6B7280;
                text-transform: uppercase;
                letter-spacing: 0.05em;
                margin-bottom: 8px;
            }}
            .kpi-value {{
                font-family: 'Plus Jakarta Sans', sans-serif;
                font-size: 26px;
                font-weight: 700;
                color: #111827;
                line-height: 1;
            }}
            
            /* Inline Alert Micro-banners */
            .status-pill-green {{
                background-color: #F0FDF4;
                color: #166534;
                padding: 4px 10px;
                border-radius: 9999px;
                font-size: 12px;
                font-weight: 600;
                display: inline-block;
            }}
        </style>
    """, unsafe_allow_html=True)

    # ==========================================
    # 📥 CORE DATA ADAPTER & CLEANING LAYER
    # ==========================================
    raw_loans = get_cached_data("loans")
    raw_borrowers = get_cached_data("borrowers")

    if raw_loans is None or raw_loans.empty:
        st.info("📅 Workspace clean. There are currently no loan records present on your corporate network.")
        return

    # Normalize Column Naming Layouts safely
    raw_loans.columns = raw_loans.columns.str.strip().str.lower().str.replace(" ", "_")
    
    # Apply tenant strict data boundaries
    if "tenant_id" in raw_loans.columns:
        loans_df = raw_loans[raw_loans["tenant_id"].astype(str) == str(current_tenant)].copy()
    else:
        loans_df = raw_loans.copy()

    if loans_df.empty:
        st.info("💡 No loans linked to this workspace registration profiles yet.")
        return

    # Dynamic Borrower Identity Linking
    if raw_borrowers is not None and not raw_borrowers.empty:
        raw_borrowers.columns = raw_borrowers.columns.str.strip().str.lower().str.replace(" ", "_")
        borrowers_df = raw_borrowers if "tenant_id" not in raw_borrowers.columns else raw_borrowers[raw_borrowers["tenant_id"].astype(str) == str(current_tenant)]
        bor_map = dict(zip(borrowers_df["id"].astype(str), borrowers_df["name"]))
    else:
        bor_map = {}

    if "borrower" not in loans_df.columns:
        loans_df["borrower"] = loans_df["borrower_id"].astype(str).map(bor_map).fillna("Unknown Profile") if "borrower_id" in loans_df.columns else "Unknown Profile"

    # ID Mapping & De-duplication Sanitizer
    def clean_id_str(val):
        if pd.isna(val): return ""
        s = str(val).strip()
        if s.endswith(".0"): return s[:-2]
        return s

    loans_df["id_clean"] = loans_df["id"].apply(clean_id_str)
    if "loan_id_label" not in loans_df.columns:
        loans_df["loan_id_label"] = loans_df["id_clean"]
    loans_df["loan_id_label"] = loans_df["loan_id_label"].fillna(loans_df["id_clean"]).astype(str)

    # Parse and safely align structural financial numbers & chronological deadlines
    date_col = "due_date" if "due_date" in loans_df.columns else "end_date"
    loans_df[date_col] = pd.to_datetime(loans_df[date_col], errors="coerce")
    loans_df["balance"] = pd.to_numeric(loans_df.get("balance", 0), errors="coerce").fillna(0.0)
    loans_df["total_repayable"] = pd.to_numeric(loans_df.get("total_repayable", loans_df.get("amount", 0)), errors="coerce").fillna(0.0)

    # Normalize structural times to absolute current local date tracking matrices
    today = pd.Timestamp.today().normalize()

    # ==========================================
    # 🔄 MULTI-CYCLE DEDUPLICATION & COMPILATION
    # ==========================================
    # Trace through sequential changes to extract the latest single version of each loan account
    seq_col = "cycle" if "cycle" in loans_df.columns else "id"
    active_loans = (
        loans_df.sort_values(by=[seq_col])
        .groupby("loan_id_label", as_index=False)
        .tail(1)
    ).copy()

    # Strip out any permanently closed asset structures or zero balances
    active_loans = active_loans[~active_loans["status"].astype(str).str.upper().isin(["PAID", "CLEARED", "CLOSED"])]
    active_loans = active_loans[active_loans["balance"].apply(lambda x: abs(x) > 5.0)]

    # Dynamic Frame Headers
    st.markdown("<h1 class='main-title'>📅 Recovery Operations Control</h1>", unsafe_allow_html=True)
    st.markdown("<p class='subtitle'>Real-time tracking of maturity pipelines, workload distributions, and exposure schedules.</p>", unsafe_allow_html=True)

    if active_loans.empty:
        st.markdown("""
            <div style="background-color: #F0FDF4; border: 1px solid #DCFCE7; padding: 24px; border-radius: 16px; text-align: center; margin-top:20px;">
                <span class="status-pill-green">All Clear</span>
                <h4 style="color:#166534; font-family:'Plus Jakarta Sans'; margin: 12px 0 4px 0; font-weight:700;">Portfolio Fully Settled</h4>
                <p style="color:#1F663C; font-family:'Plus Jakarta Sans'; font-size:14px; margin:0;">No loans currently hold outstanding maturity metrics or past-due balances for this lifecycle period.</p>
            </div>
        """, unsafe_allow_html=True)
        return

    # ==========================================
    # 📊 EXECUTIVE LEVEL KPI OVERVIEWS
    # ==========================================
    due_today_df = active_loans[active_loans[date_col].dt.date == today.date()]
    upcoming_df = active_loans[
        (active_loans[date_col].dt.date > today.date()) & 
        (active_loans[date_col].dt.date <= (today + pd.Timedelta(days=7)).date())
    ]
    overdue_df = active_loans[active_loans[date_col].dt.date < today.date()]

    kpi1, kpi2, kpi3 = st.columns(3)
    
    with kpi1:
        st.markdown(f"""
            <div class="kpi-container" style="border-top: 4px solid {brand_color};">
                <div class="kpi-label">Maturing Today</div>
                <div class="kpi-value">{len(due_today_df)} <span style="font-size:14px; font-weight:500; color:#6B7280;">Accounts</span></div>
            </div>
        """, unsafe_allow_html=True)
        
    with kpi2:
        st.markdown(f"""
            <div class="kpi-container" style="border-top: 4px solid #3B82F6;">
                <div class="kpi-label">Upcoming Window (7 Days)</div>
                <div class="kpi-value">{len(upcoming_df)} <span style="font-size:14px; font-weight:500; color:#6B7280;">Accounts</span></div>
            </div>
        """, unsafe_allow_html=True)
        
    with kpi3:
        # Accent changes dynamically based on active system overdue risk exposure limits
        accent = "#EF4444" if not overdue_df.empty else "#10B981"
        st.markdown(f"""
            <div class="kpi-container" style="border-top: 4px solid {accent};">
                <div class="kpi-label">Outstanding Overdue Backlog</div>
                <div class="kpi-value" style="color: {accent};">{len(overdue_df)} <span style="font-size:14px; font-weight:500; color:#6B7280;">Accounts</span></div>
            </div>
        """, unsafe_allow_html=True)

    st.markdown("<br><hr style='border:0; border-top:1px solid #E5E7EB; margin:10px 0 25px 0;' />", unsafe_allow_html=True)

    # ==========================================
    # 🗓️ VISUAL CALENDAR COMPONENT TARGET LAYER
    # ==========================================
    st.markdown("<h3 style='font-family:\"Plus Jakarta Sans\"; font-weight:700; font-size:18px; color:#111827; margin-bottom:12px;'>🗓️ Interactive Operations Timeline</h3>", unsafe_allow_html=True)
    
    events_list = []
    for _, item in active_loans.iterrows():
        if pd.notna(item[date_col]):
            is_past = item[date_col].date() < today.date()
            color_code = "#EF4444" if is_past else "#2B3F87"
            
            events_list.append({
                "title": f"UGX {float(item['balance']):,.0f} - {item['borrower']}",
                "start": item[date_col].strftime("%Y-%m-%d"),
                "end": item[date_col].strftime("%Y-%m-%d"),
                "color": color_code,
                "allDay": True,
            })

    options_config = {
        "headerToolbar": {"left": "prev,next today", "center": "title", "right": "dayGridMonth,timeGridWeek"},
        "initialView": "dayGridMonth",
        "selectable": True,
        "editable": False
    }

    try:
        from streamlit_calendar import st_calendar
        st_calendar(events=events_list, options=options_config, key="modern_collection_calendar")
    except ImportError:
        # Beautiful UI fallback element block if plugin execution environment variations appear
        st.info("ℹ️ Advanced visual calendar package module not registered. Using data stream grids directly.")

    st.markdown("<br>", unsafe_allow_html=True)

    # ==========================================
    # ⚡ OPERATIONAL SUBDIVISIONS & SCHEDULERS
    # ==========================================
    layout_left, layout_right = st.columns([4, 5])

    with layout_left:
        st.markdown("<h3 style='font-family:\"Plus Jakarta Sans\"; font-weight:700; font-size:18px; color:#111827; margin-bottom:15px;'>🎯 Due Action Plan (Today)</h3>", unsafe_allow_html=True)
        
        if due_today_df.empty:
            st.markdown("""
                <div style="background-color: #F9FAFB; border: 1px dashed #E5E7EB; padding: 20px; border-radius: 12px; text-align: center;">
                    <p style="color:#6B7280; font-family:'Plus Jakarta Sans'; font-size:13px; margin:0;">✨ No collection deadlines slated for today's date context.</p>
                </div>
            """, unsafe_allow_html=True)
        else:
            display_today = due_today_df.copy()
            display_today["Short ID"] = display_today.apply(lambda r: f"#{r['loan_id_label']}", axis=1)
            
            st.dataframe(
                display_today[["Short ID", "borrower", "balance"]].rename(columns={
                    "Short ID": "Loan ID",
                    "borrower": "Client Name",
                    "balance": "Balance Owed"
                }),
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Balance Owed": st.column_config.NumberColumn(format="UGX %,.0f")
                }
            )

    with layout_right:
        st.markdown("<h3 style='font-family:\"Plus Jakarta Sans\"; font-weight:700; font-size:18px; color:#111827; margin-bottom:15px;'>🔮 30-Day Workload Projection</h3>", unsafe_allow_html=True)
        
        current_month_df = active_loans[active_loans[date_col].dt.month == today.month]
        month_total_expected = current_month_df["balance"].sum()
        
        col_m1, col_m2 = st.columns(2)
        with col_m1:
            st.metric(
                label="Target Pipeline Exposure", 
                value=f"UGX {month_total_expected:,.0f}", 
                delta=f"{len(current_month_df)} Accounts Active",
                delta_color="normal"
            )
        with col_m2:
            avg_ticket = current_month_df["balance"].mean() if not current_month_df.empty else 0.0
            st.metric(
                label="Mean Account Exposure", 
                value=f"UGX {avg_ticket:,.0f}",
                delta="Rolling Month Average",
                delta_color="off"
            )

    # ==========================================
    # 🔴 AUDIT-READY RISK RECOVERY LEDGER
    # ==========================================
    st.markdown("<br><h3 style='font-family:\"Plus Jakarta Sans\"; font-weight:700; font-size:18px; color:#EF4444; margin-bottom:15px;'>🔴 High Risk Past-Due Recoveries</h3>", unsafe_allow_html=True)
    
    if not overdue_df.empty:
        overdue_display = overdue_df.copy()
        overdue_display["days_late"] = (today - overdue_display[date_col]).dt.days
        overdue_display["Arrears Duration"] = overdue_display["days_late"].apply(lambda x: f"⚠️ {x} Days Late")
        overdue_display["Short ID"] = overdue_display.apply(lambda r: f"#{r['loan_id_label']}", axis=1)

        st.dataframe(
            overdue_display.sort_values("days_late", ascending=False)[
                ["Short ID", "borrower", "Arrears Duration", "balance", "status"]
            ].rename(columns={
                "Short ID": "Loan Account",
                "borrower": "Client Identity Name",
                "balance": "Outstanding Exposure (UGX)",
                "status": "System Base Status"
            }),
            use_container_width=True,
            hide_index=True,
            column_config={
                "Outstanding Exposure (UGX)": st.column_config.NumberColumn(format="%,.0f"),
                "System Base Status": st.column_config.TextColumn(help="Raw lifecycle phase flag recorded in system tables")
            }
        )
    else:
        st.markdown("""
            <div style="background-color: #F0FDF4; border: 1px dashed #BBF7D0; padding: 16px; border-radius: 12px; margin-bottom: 20px;">
                <p style="color:#15803D; font-family:'Plus Jakarta Sans'; font-size:13px; margin:0; font-weight:500;">
                    ✨ Clean Portfolio Pipeline Status: No past-due accounts currently require active collection management workflows.
                </p>
            </div>
        """, unsafe_allow_html=True)
# ==========================================================
# 🛡️ COLLATERAL MANAGEMENT ENGINE
# ==========================================================
import streamlit as st
import pandas as pd
from datetime import datetime

def show_collateral():
    """
    Manages asset security, verification tracking, and real-time status management
    while respecting multi-tenant boundaries.
    """
    brand_color = st.session_state.get("theme_color", "#2B3F87")
    current_tenant = st.session_state.get('tenant_id')

    if not current_tenant:
        st.error("🔐 Session expired. Please log in.")
        st.stop()

    st.markdown(f"<h2 style='color: {brand_color}; margin-bottom: 20px;'>🛡️ Collateral & Security Assets</h2>", unsafe_allow_html=True)

    # ==============================
    # 📦 DATA FETCH & UNIFICATION
    # ==============================
    collateral_df = get_data("collateral") 
    loans_df = get_data("loans")
    borrowers_df = get_data("borrowers")  # Added to ensure fallback relational mapping works safely

    # Protect against missing data collections
    if loans_df is None: loans_df = pd.DataFrame()
    if collateral_df is None: collateral_df = pd.DataFrame()
    if borrowers_df is None: borrowers_df = pd.DataFrame()

    # Tenant Isolation Filter
    if not loans_df.empty and "tenant_id" in loans_df.columns:
        loans_df = loans_df[loans_df["tenant_id"].astype(str) == str(current_tenant)].copy()
    if not collateral_df.empty and "tenant_id" in collateral_df.columns:
        collateral_df = collateral_df[collateral_df["tenant_id"].astype(str) == str(current_tenant)].copy()

    # --- RELATIONAL BORROWER NAME LOOKUP ---
    # Safely creates name mappings without breaking case conventions of the tables
    name_map = {}
    if not borrowers_df.empty:
        b_id_col = next((c for c in borrowers_df.columns if c.lower() in ['id', 'borrower_id']), None)
        b_nm_col = next((c for c in borrowers_df.columns if c.lower() in ['name', 'borrower_name', 'client_name']), None)
        if b_id_col and b_nm_col:
            name_map = dict(zip(borrowers_df[b_id_col].astype(str), borrowers_df[b_nm_col]))

    # Add descriptive borrower labels inline to loans dataframe safely
    if not loans_df.empty:
        loans_df['id_str'] = loans_df['id'].astype(str)
        if 'borrower_id' in loans_df.columns:
            loans_df['borrower_name_resolved'] = loans_df['borrower_id'].astype(str).map(name_map)
        else:
            # Fallback inline string lookup if tables are denormalized
            match_col = next((c for c in loans_df.columns if c.lower() in ['borrower', 'client', 'borrower_name']), None)
            loans_df['borrower_name_resolved'] = loans_df[match_col] if match_col else "Unknown Borrower"
        
        loans_df['borrower_name_resolved'] = loans_df['borrower_name_resolved'].fillna("Unknown Borrower")
        
        # Filter down to operational loans matching dynamic context
        status_col = next((c for c in loans_df.columns if c.lower() == 'status'), None)
        if status_col:
            active_statuses = ["ACTIVE", "OVERDUE", "PENDING", "BCF"]
            available_loans = loans_df[loans_df[status_col].astype(str).str.upper().isin(active_statuses)].copy()
        else:
            available_loans = loans_df.copy()
    else:
        available_loans = pd.DataFrame()

    # --- TABS CONTAINER ---
    tab_reg, tab_view = st.tabs(["➕ Register Asset", "📋 Inventory & Status"])

    # ==============================
    # ➕ TAB 1: REGISTER ASSET
    # ==============================
    with tab_reg:
        if available_loans.empty:
            st.info("ℹ️ No active context loans found to connect security collateral onto.")
        else:
            with st.form("collateral_reg_form", clear_on_submit=True):
                st.write("### Link Asset to Loan File")
                c1, c2 = st.columns(2)

                # Generate clean dropdown tracking options
                loan_map = {}
                for _, row in available_loans.iterrows():
                    loan_id = str(row['id'])
                    b_name = str(row['borrower_name_resolved'])
                    if b_name.lower() in ['nan', 'none', '']: b_name = "Unknown Customer"
                    
                    ref_label = row.get('loan_id_label', loan_id[:8])
                    principal_amt = row.get('principal', row.get('amount', 0))
                    try:
                        amt_fmt = f"UGX {float(principal_amt):,.0f}"
                    except Exception:
                        amt_fmt = f"UGX {principal_amt}"

                    loan_map[loan_id] = f"{b_name} | {amt_fmt} (Ref: {ref_label})"

                selected_loan_id = c1.selectbox(
                    "Select Target Loan Profile",
                    options=list(loan_map.keys()),
                    format_func=lambda x: loan_map.get(x, "Select Profile")
                )

                asset_type = c2.selectbox(
                    "Asset Classification Type",
                    ["Logbook (Car)", "Land Title", "Electronics", "House Deed", "Business Stock", "Other"]
                )

                desc = st.text_input("Detailed Description (Asset Serial ID, Plate Number, Plot Location)")
                est_value = st.number_input("Estimated Asset Market Value (UGX)", min_value=0, step=100000)
                
                st.markdown("<br>", unsafe_allow_html=True)
                uploaded_photo = st.file_uploader("Upload Verification Documentation/Photo Asset", type=["jpg", "png", "jpeg"])

                submit_save = st.form_submit_button("💾 Save Asset Registration Record", use_container_width=True)

            if submit_save:
                if not desc or est_value <= 0:
                    st.error("❌ Form Incomplete. Provide accurate structural evaluations and descriptions.")
                else:
                    try:
                        # Safely compute borrower back-reference assignment details
                        full_label = loan_map[selected_loan_id]
                        borrower_for_db = full_label.split(" | ")[0]

                        new_asset = pd.DataFrame([{
                            "loan_id": selected_loan_id,
                            "tenant_id": str(current_tenant),
                            "borrower": borrower_for_db,
                            "type": asset_type,
                            "description": desc,
                            "value": float(est_value),
                            "status": "In Custody",
                            "date_added": datetime.now().strftime("%Y-%m-%d")
                        }])

                        if save_data_saas("collateral", new_asset):
                            st.success(f"✅ Asset registry successfully locked down for {borrower_for_db}!")
                            st.cache_data.clear()
                            st.rerun()
                    except Exception as e:
                        st.error(f"❌ Security transaction failed to register: {e}")

    # ==============================
    # 📋 TAB 2: INVENTORY & LEDGER
    # ==============================
    with tab_view:
        if collateral_df.empty:
            st.info("💡 No securitized collateral structures found inside this workspace index.")
        else:
            # Clean and normalize asset value matrices safely
            val_col = next((c for c in collateral_df.columns if c.lower() == 'value'), 'value')
            collateral_df[val_col] = pd.to_numeric(collateral_df[val_col], errors="coerce").fillna(0)
            
            stat_col = next((c for c in collateral_df.columns if c.lower() == 'status'), 'status')
            
            # --- KPI DASHBOARD CARDS ---
            total_value = collateral_df[val_col].sum()
            held_count = len(collateral_df[collateral_df[stat_col].astype(str).str.upper() == "IN CUSTODY"])

            m1, m2 = st.columns(2)
            m1.metric("Securitized Inventory Value", f"UGX {total_value:,.0f}")
            m2.metric("Active Assets in Custody", held_count)
            st.markdown("---")

            # --- DYNAMIC FILTERS LAYER ---
            col1, col2 = st.columns(2)
            unique_statuses = sorted(collateral_df[stat_col].dropna().unique().tolist())
            status_filter = col1.selectbox("Filter Inventory Status", ["All Asset Records"] + unique_statuses)
            search_query = col2.text_input("🔍 Search Keyword / Borrower Reference").lower()

            df_filtered = collateral_df.copy()
            if status_filter != "All Asset Records":
                df_filtered = df_filtered[df_filtered[stat_col] == status_filter]

            if search_query:
                b_match = next((c for c in df_filtered.columns if c.lower() == 'borrower'), 'borrower')
                d_match = next((c for c in df_filtered.columns if c.lower() == 'description'), 'description')
                
                df_filtered = df_filtered[
                    df_filtered[b_match].astype(str).str.lower().str.contains(search_query, na=False) |
                    df_filtered[d_match].astype(str).str.lower().str.contains(search_query, na=False)
                ]

            # --- RENDER UNIFORM JAVASCRIPT DATAFRAME ---
            st.markdown("### Securitized Assets Ledger")
            
            display_ledger = df_filtered.copy()
            # Remap columns neatly to standard clean headings
            col_renames = {
                "date_added": "Date Registered",
                "borrower": "Borrower Profile",
                "type": "Asset Type",
                "description": "Asset Description",
                "value": "Value (UGX)",
                "status": "Tracking Status"
            }
            # Handle minor schema variances gracefully
            display_ledger = display_ledger.rename(columns={k: v for k, v in col_renames.items() if k in display_ledger.columns})
            
            # Formats display currencies seamlessly 
            tgt_val_col = "Value (UGX)"
            if tgt_val_col in display_ledger.columns:
                display_ledger[tgt_val_col] = display_ledger[tgt_val_col].apply(lambda x: f"{float(x):,.0f}" if pd.notna(x) else "0")

            cols_to_show = [v for v in col_renames.values() if v in display_ledger.columns]
            
            st.dataframe(
                display_ledger[cols_to_show],
                use_container_width=True,
                hide_index=True
            )
            st.markdown("---")

            # ==============================
            # ⚙️ SECURE LIFECYCLE MANAGEMENT
            # ==============================
            st.markdown("### 🛠️ Strategic Asset Inspection & Audits")
            
            if df_filtered.empty:
                st.warning("No records match configuration queries.")
            else:
                b_lbl = next((c for c in df_filtered.columns if c.lower() == 'borrower'), 'borrower')
                d_lbl = next((c for c in df_filtered.columns if c.lower() == 'description'), 'description')
                
                df_filtered["ui_label"] = df_filtered[b_lbl].astype(str) + " — " + df_filtered[d_lbl].astype(str)
                selected_label = st.selectbox("Choose Target Asset Inventory File", df_filtered["ui_label"].tolist())

                selected_row = df_filtered[df_filtered["ui_label"] == selected_label].iloc[0]
                asset_id = selected_row["id"]

                # --- PHOTO EVIDENCE STORAGE ---
                st.markdown("#### 📸 Physical Asset Evidence")
                asset_photo = selected_row.get("photo", selected_row.get("image_url", None))
                if asset_photo and pd.notna(asset_photo):
                    st.image(asset_photo, caption=str(selected_row[d_lbl]), use_container_width=True)
                else:
                    st.info("No visual photo documentation uploaded against this asset record.")
                
                st.markdown("<br>", unsafe_allow_html=True)

                # --- LIFECYCLE OPERATION PANEL ---
                st.markdown("#### 🔄 Asset State Custody Release")
                status_options = ["In Custody", "Released", "Disposed (Auctioned)", "Held for Pickup"]
                
                # Fixed layout alignment split to prevent text wrapping on operational buttons
                col_stat, col_btn = st.columns([2, 1])
                
                current_status_val = str(selected_row[stat_col])
                default_idx = status_options.index(current_status_val) if current_status_val in status_options else 0

                new_status = col_stat.selectbox(
                    "Transition Status State",
                    status_options,
                    index=default_idx,
                    label_visibility="collapsed"
                )

                # Standard button layout prevents regressions and aligns cleanly
                if col_btn.button("Update Asset State", use_container_width=True):
                    update_row = pd.DataFrame([{
                        "id": asset_id,
                        "status": new_status,
                        "tenant_id": str(current_tenant)
                    }])

                    if save_data_saas("collateral", update_row):
                        st.success(f"✅ Custody updated to '{new_status}' successfully!")
                        st.cache_data.clear()
                        st.rerun()
import streamlit as st
import pandas as pd
import uuid
import plotly.express as px
from datetime import datetime

def show_expenses():
    """
    Manages internal operational company costs, budgeting lifecycles, 
    and multi-tenant spending trends.
    """
    brand_color = st.session_state.get("theme_color", "#2B3F87")
    current_tenant = st.session_state.get('tenant_id')

    if not current_tenant:
        st.error("🔐 Session expired. Please log in.")
        st.stop()

    st.markdown(f"<h2 style='color: {brand_color}; margin-bottom: 20px;'>📁 Expense Management</h2>", unsafe_allow_html=True)

    def get_fy_label(date_val):
        try:
            dt = pd.to_datetime(date_val)
            if pd.isna(dt): return "Unknown FY"
            return f"FY{dt.year}-{dt.year+1}" if dt.month >= 7 else f"FY{dt.year-1}-{dt.year}"
        except:
            return "Unknown FY"

    # ==============================
    # 📥 DATA INGESTION & BOUNDARY LAYERS
    # ==============================
    try:
        raw_all_df = get_cached_data("expenses")
    except:
        raw_all_df = pd.DataFrame()

    if raw_all_df is not None and not raw_all_df.empty:
        raw_all_df.columns = raw_all_df.columns.str.lower().str.strip()
        raw_all_df["id"] = raw_all_df["id"].astype(str)
        raw_all_df["amount"] = pd.to_numeric(raw_all_df["amount"], errors="coerce").fillna(0.0)
        
        if "tenant_id" in raw_all_df.columns:
            df = raw_all_df[raw_all_df["tenant_id"].astype(str) == str(current_tenant)].copy()
        else:
            df = raw_all_df.copy()
            
        if not df.empty and "payment_date" in df.columns:
            df["financial_year"] = df["payment_date"].apply(get_fy_label)
        else:
            df["financial_year"] = "Unknown FY"
    else:
        df = pd.DataFrame(columns=["id","category","amount","date","description","payment_date","receipt_no","tenant_id","financial_year"])
        raw_all_df = df.copy()

    EXPENSE_CATS = ["Rent","Insurance","Utilities","Salaries","Licence Expenses","Marketing","Office Expenses","Operating Expenses","Fuel and Motor Vehicle","Taxes","Corporate Social Responsibilities","Other"]

    tab_add, tab_view, tab_manage = st.tabs([
        "➕ Record Expense", "📊 Spending Analysis", "⚙️ Manage Records"
    ])

    # ==============================
    # ➕ TAB 1: RECORD EXPENSE
    # ==============================
    with tab_add:
        with st.form("add_expense_form", clear_on_submit=True):
            st.write("### Log Operational Cost")
            c1, c2 = st.columns(2)

            category = c1.selectbox("Expense Category", EXPENSE_CATS)
            amount = c2.number_input("Amount (UGX)", min_value=0.0, step=10000.0)  # Typed to floats
            desc = st.text_input("Transaction Description / Payee")

            c3, c4 = st.columns(2)
            p_date = c3.date_input("Payment Date", value=datetime.now())
            receipt_no = c4.text_input("Receipt / Invoice Number Reference")

            submit_new = st.form_submit_button("🚀 Save Expense Record", use_container_width=True)

            if submit_new:
                if amount > 0 and desc:
                    try:
                        formatted_d = p_date.strftime("%Y-%m-%d")

                        new_record = pd.DataFrame([{
                            "id": str(uuid.uuid4()),
                            "category": category,
                            "amount": float(amount),
                            "date": formatted_d,
                            "description": desc,
                            "payment_date": formatted_d,
                            "receipt_no": receipt_no,
                            "tenant_id": str(current_tenant)
                        }])

                        updated_global_df = pd.concat([raw_all_df, new_record], ignore_index=True)
                        if "financial_year" in updated_global_df.columns:
                            updated_global_df = updated_global_df.drop(columns=["financial_year"], errors="ignore")

                        if save_data("expenses", updated_global_df):
                            st.success(f"✅ Expense logged successfully for {formatted_d}!")
                            st.cache_data.clear()
                            st.rerun()
                    except Exception as e:
                        st.error(f"🚨 Operational save failed: {e}")
                else:
                    st.warning("⚠️ Valid execution amount and asset descriptions are required.")

    # ==============================
    # 📊 TAB 2: SPENDING ANALYSIS
    # ==============================
    with tab_view:
        if df.empty:
            st.info("💡 No transactional outlays registered yet inside this profile.")
        else:
            fys = sorted(df["financial_year"].unique(), reverse=True)
            fy = st.selectbox("📅 Structural Scope Window (Financial Year)", ["All Time Target"] + fys)

            view_df = df if fy == "All Time Target" else df[df["financial_year"] == fy]
            total_outflow = view_df["amount"].sum()

            st.markdown(f"""
            <div style="background-color:white; padding:20px; border-radius:12px; border-left:6px solid #FF4B4B; box-shadow:0 2px 8px rgba(0,0,0,0.04); margin-bottom: 20px;">
                <p style="margin:0; font-size:11px; color:#6B7280; font-weight:600; text-transform:uppercase;">Aggregate Operational Outflow ({fy})</p>
                <h2 style="margin:5px 0 0 0; color:#1F2937; font-weight:700;">UGX {total_outflow:,.0f}</h2>
            </div>
            """, unsafe_allow_html=True)

            col1, col2 = st.columns([2, 1])
            with col1:
                chart_data = view_df.groupby("category")["amount"].sum().reset_index()
                fig = px.pie(
                    chart_data, names="category", values="amount",
                    hole=0.42, title="Cost Contribution Share",
                    color_discrete_sequence=px.colors.qualitative.Safe
                )
                fig.update_layout(margin=dict(t=40, b=10, l=10, r=10), legend=dict(orientation="h", y=-0.1))
                st.plotly_chart(fig, use_container_width=True)

            with col2:
                st.markdown("#### Annual Trajectory")
                fy_summary = df.groupby("financial_year")["amount"].sum().reset_index().rename(
                    columns={"financial_year": "Fin Year", "amount": "Total (UGX)"}
                )
                fy_summary["Total (UGX)"] = fy_summary["Total (UGX)"].apply(lambda x: f"{x:,.0f}")
                st.dataframe(fy_summary, use_container_width=True, hide_index=True)

            st.markdown("### 📋 Expense Ledger Records")
            
            filter_c1, filter_c2 = st.columns(2)
            cat_options = ["All Categories"] + sorted(view_df["category"].dropna().unique().tolist())
            selected_cat = filter_c1.selectbox("Category Filter Selection", cat_options)

            min_val = float(view_df["amount"].min()) if not view_df.empty else 0.0
            max_val = float(view_df["amount"].max()) if not view_df.empty else 100000.0
            if min_val == max_val: max_val += 1.0
            
            amt_range = filter_c2.slider("Transaction Cost Range Threshold", min_val, max_val, (min_val, max_val))

            processed_df = view_df.copy()
            if selected_cat != "All Categories":
                processed_df = processed_df[processed_df["category"] == selected_cat]
            
            processed_df = processed_df[
                (processed_df["amount"] >= amt_range[0]) & 
                (processed_df["amount"] <= amt_range[1])
            ]

            if processed_df.empty:
                st.warning("No records align with current query filters.")
            else:
                processed_df = processed_df.sort_values("payment_date", ascending=False)
                display_ledger = processed_df[["payment_date", "category", "description", "amount", "receipt_no"]].copy()
                display_ledger.columns = ["Date Paid", "Category Grouping", "Description Detail", "Amount (UGX)", "Ref / Invoice #"]
                display_ledger["Amount (UGX)"] = display_ledger["Amount (UGX)"].apply(lambda x: f"{x:,.0f}")
                
                st.dataframe(
                    display_ledger,
                    use_container_width=True,
                    hide_index=True
                )

    # ==============================
    # ⚙️ TAB 3: MANAGE RECORDS
    # ==============================
    with tab_manage:
        st.markdown("### 🛠️ Record Maintenance Engine")

        if df.empty:
            st.info("No active expense records discovered to modify.")
        else:
            df["selector_label"] = df.apply(
                lambda r: f"{r['payment_date']} | {r['category']} | UGX {r['amount']:,.0f} — {r['description'][:15]}...", axis=1
            )

            record_map = {row["selector_label"]: row for _, row in df.iterrows()}
            selected_label = st.selectbox("Choose Target Transaction Log Entry", list(record_map.keys()))

            if selected_label:
                target_record = record_map[selected_label]
                
                with st.form("edit_expense_form"):
                    st.write(f"**Modifying Reference Key**: `{target_record['id'][:8]}`")
                    
                    # 🛡️ THE FIX: Match float value type with an explicit float step representation
                    new_amt = st.number_input("Update Value Allocation (UGX)", value=float(target_record['amount']), step=5000.0)
                    new_desc = st.text_input("Modify Description Label Details", value=target_record['description'])
                    
                    mod_c1, mod_c2 = st.columns(2)
                    save_btn = mod_c1.form_submit_button("💾 Save Matrix Modifications", use_container_width=True)
                    delete_btn = mod_c2.form_submit_button("🗑️ Purge Log Record", use_container_width=True)

                    if save_btn:
                        raw_all_df.loc[raw_all_df["id"] == target_record["id"], ["amount", "description"]] = [float(new_amt), new_desc]
                        if "financial_year" in raw_all_df.columns:
                            raw_all_df = raw_all_df.drop(columns=["financial_year"], errors="ignore")

                        if save_data("expenses", raw_all_df):
                            st.success("✅ Log tracking changes updated successfully!")
                            st.cache_data.clear()
                            st.rerun()

                    if delete_btn:
                        clean_global_df = raw_all_df[raw_all_df["id"] != str(target_record["id"])].copy()
                        if "financial_year" in clean_global_df.columns:
                            clean_global_df = clean_global_df.drop(columns=["financial_year"], errors="ignore")

                        if save_data("expenses", clean_global_df):
                            st.warning("🗑️ Expense item purged from system logs.")
                            st.cache_data.clear()
                            st.rerun()


    
import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
# Ensure proper ReportLab styling imports are active
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def generate_pdf_statement(client_name, loans_df, payments_df):
    """
    Compiles a clean financial ledger PDF statement safely grouped by unique 
    Loan Ref ID, pulling only the true original disbursement and subsequent actions.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    
    cell_style = ParagraphStyle('LedgerCell', parent=styles['Normal'], fontSize=8, leading=10)
    cell_bold = ParagraphStyle('LedgerCellBold', parent=styles['Normal'], fontSize=8, leading=10, fontName='Helvetica-Bold')

    elements = []
    company_name = st.session_state.get('company_name', 'ZOE CONSULTS SMC LIMITED').upper()
    elements.append(Paragraph(f"<b>{company_name}</b>", styles["Title"]))
    elements.append(Paragraph(f"<b>Client Statement:</b> {client_name}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Statement Date:</b> {datetime.now().strftime('%d %b %Y')}", styles["Normal"]))
    elements.append(Spacer(1, 20))
    
    grand_total = 0.0
    
    def clean_id_str(val):
        if pd.isna(val): return ""
        s = str(val).strip()
        if s.endswith(".0"): s = s[:-2]
        return s

    processed_payments = payments_df.copy() if payments_df is not None else pd.DataFrame()
    if not processed_payments.empty and "loan_id" in processed_payments.columns:
        processed_payments["loan_id"] = processed_payments["loan_id"].apply(clean_id_str)

    working_loans = loans_df.copy()
    working_loans["id_clean"] = working_loans["id"].apply(clean_id_str)
    
    if "loan_id_label" not in working_loans.columns:
        working_loans["loan_id_label"] = working_loans["id_clean"]
    working_loans["loan_id_label"] = working_loans["loan_id_label"].fillna(working_loans["id_clean"]).astype(str)

    grouped_loan_labels = sorted(working_loans["loan_id_label"].unique())

    for display_id in grouped_loan_labels:
        sub_loans = working_loans[working_loans["loan_id_label"] == display_id]
        
        # 🔄 Sort by cycle or sequence key to trace the true initial entry
        seq_col = "cycle" if "cycle" in sub_loans.columns else "id"
        sub_loans = working_loans[
            working_loans["loan_id_label"] == display_id
        ].copy()
        
        date_col = (
            "start_date"
            if "start_date" in sub_loans.columns
            else "created_at"
        )
        
        sub_loans[date_col] = pd.to_datetime(
            sub_loans[date_col],
            errors="coerce"
        )
        
        sub_loans = sub_loans.sort_values(
            by=[date_col]
        )
        
        sub_loans = sub_loans.drop_duplicates(
            subset=[date_col, "balance"],
            keep="last"
        )
        
        origin_loan = sub_loans.iloc[0]
        
        elements.append(Paragraph(f"<b>Loan Account Ref: {display_id}</b>", styles["Heading3"]))
        
        data = [[
            Paragraph("<b>Cycle Date</b>", cell_bold),
            Paragraph("<b>Cycle</b>", cell_bold),
            Paragraph("<b>Opening Balance</b>", cell_bold),
            Paragraph("<b>Amount Due</b>", cell_bold),
            Paragraph("<b>Amount Paid</b>", cell_bold),
            Paragraph("<b>Balance</b>", cell_bold),
        ]]
        
        for _, row in sub_loans.iterrows():
        
            cycle_date = str(
                row.get(
                    "start_date",
                    row.get("created_at", "")
                )
            )[:10]
        
            opening_balance = float(
            row.get(
                "opening_balance",
                row.get(
                    "principal",
                    0
                )
            )
        )
        
            amount_due = float(
                row.get(
                    "total_due",
                    row.get("amount_due", 0)
                )
            )
        
            amount_paid = float(
                row.get(
                    "amount_paid",
                    row.get("paid", 0)
                )
            )
        
            balance = float(
                row.get(
                    "balance",
                    0
                )
            )
        
            data.append([
                Paragraph(cycle_date, cell_style),
                Paragraph(str(row.get("cycle", "")), cell_style),
                Paragraph(f"{opening_balance:,.0f}", cell_style),
                Paragraph(f"{amount_due:,.0f}", cell_style),
                Paragraph(f"{amount_paid:,.0f}", cell_style),
                Paragraph(f"{balance:,.0f}", cell_style),
            ])
        
        final_balance = float(sub_loans.iloc[-1].get("balance", 0))
        
        grand_total += abs(final_balance)

        table = Table(
            data,
            repeatRows=1,
            colWidths=[65, 40, 85, 85, 85, 85]
        )
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#89CFF0")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D6EAF8")),
        
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
                colors.white,
                colors.HexColor("#F8FCFF")
            ]),
        
            ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
        
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]))
        elements.append(table)
        status_value = str(
            sub_loans.iloc[-1].get(
                "status",
                "OUTSTANDING"
            )
        )
        
        final_balance = abs(
            float(
                sub_loans.iloc[-1].get(
                    "balance",
                    0
                )
            )
        )
        
        if final_balance <= 1:
            status_value = "CLEARED"
        else:
            status_value = "PENDING"
        
        elements.append(
            Spacer(1, 5)
        )
        
        loan_style = ParagraphStyle(
            "LoanHeading",
            parent=styles["Heading3"],
            textColor=colors.HexColor("#89CFF0")
        )
        
        elements.append(
            Paragraph(
                f"<b>Loan Account Ref: {display_id}</b>",
                loan_style
            )
        )
        elements.append(Spacer(1, 15))
        
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"<b>Aggregate Outstanding Position: {grand_total:,.0f} UGX</b>", styles["Heading2"]))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer

# ==============================
# MAIN LEDGER FUNCTION
# ==============================
def show_ledger():
    brand_color = st.session_state.get("theme_color", "#2B3F87")
    current_tenant = st.session_state.get('tenant_id')

    if not current_tenant:
        st.error("🔐 Session expired. Please log in.")
        st.stop()

    st.markdown(f"""
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');
            .ledger-header {{
                font-family: 'Inter', sans-serif;
                color: {brand_color};
                font-weight: 700;
                margin-bottom: 20px;
            }}
            .snapshot-text {{
                font-family: 'Inter', sans-serif;
                font-weight: 600;
                color: #374151;
                margin-top: 15px;
            }}
        </style>
        <h2 class='ledger-header'>📘 Master Ledger Accounts</h2>
    """, unsafe_allow_html=True)

    raw_loans = get_cached_data("loans")
    raw_payments = get_cached_data("payments")
    raw_borrowers = get_cached_data("borrowers")

    if raw_loans is None or raw_loans.empty:
        st.info("💡 Complete clear! No active loan parameters tracked on database arrays.")
        return

    raw_loans.columns = raw_loans.columns.str.strip().str.lower().str.replace(" ", "_")
    
    if "tenant_id" in raw_loans.columns:
        loans_df = raw_loans[raw_loans["tenant_id"].astype(str) == str(current_tenant)].copy()
    else:
        loans_df = raw_loans.copy()

    if loans_df.empty:
        st.info("💡 No active profile accounts registered to this company portal.")
        return

    if raw_payments is not None and not raw_payments.empty:
        raw_payments.columns = raw_payments.columns.str.strip().str.lower().str.replace(" ", "_")
        payments_df = raw_payments if "tenant_id" not in raw_payments.columns else raw_payments[raw_payments["tenant_id"].astype(str) == str(current_tenant)].copy()
    else:
        payments_df = pd.DataFrame()

    bor_map = {}
    if raw_borrowers is not None and not raw_borrowers.empty:
        raw_borrowers.columns = raw_borrowers.columns.str.strip().str.lower().str.replace(" ", "_")
        borrowers_df = raw_borrowers if "tenant_id" not in raw_borrowers.columns else raw_borrowers[raw_borrowers["tenant_id"].astype(str) == str(current_tenant)]
        bor_map = dict(zip(borrowers_df["id"].astype(str), borrowers_df["name"]))
        
    if "borrower" not in loans_df.columns:
        if "borrower_id" in loans_df.columns:
            loans_df["borrower"] = loans_df["borrower_id"].astype(str).map(bor_map).fillna("Unknown Profile")
        else:
            loans_df["borrower"] = "Unknown Profile"

    def clean_id_str(val):
        if pd.isna(val): return ""
        s = str(val).strip()
        if s.endswith(".0"): return s[:-2]
        return s

    loans_df["id_clean"] = loans_df["id"].apply(clean_id_str)
    if not payments_df.empty and "loan_id" in payments_df.columns:
        payments_df["loan_id"] = payments_df["loan_id"].apply(clean_id_str)

    if "loan_id_label" not in loans_df.columns:
        loans_df["loan_id_label"] = loans_df["id_clean"]
    loans_df["loan_id_label"] = loans_df["loan_id_label"].fillna(loans_df["id_clean"]).astype(str)

    loans_df = loans_df.sort_values(by=["loan_id_label"])

    # ==============================
    # 🎯 SELECTION INTERFACE
    # ==============================
    loan_map = {
        f"Ref: {r['loan_id_label']} — {r['borrower']}": r["loan_id_label"]
        for _, r in loans_df.iterrows()
    }
    
    selected_label = st.selectbox("🎯 Target Loan Account Filter", list(dict.fromkeys(loan_map.keys())))
    target_label = loan_map[selected_label]
    
    filtered_loans = loans_df[loans_df["loan_id_label"] == target_label]
    
    if filtered_loans.empty:
        st.error("Target loan data vector unreadable.")
        return

    # Sort chronological subsets to split initial entries vs rolling items correctly
    seq_col = "cycle" if "cycle" in filtered_loans.columns else "id"
    filtered_loans = filtered_loans.sort_values(by=[seq_col])

    # ==============================
    # 📊 CORRECTED STATEMENT PREVIEW PANEL
    # ==============================
    st.markdown("<h4 class='snapshot-text'>📑 Account Balance Breakdown</h4>", unsafe_allow_html=True)
    
    # 🔄 FIX: Isolate origin and endpoint records to get accurate, non-cumulative numbers
    origin_record = filtered_loans.iloc[0]
    latest_record = filtered_loans.iloc[-1]

    p = float(origin_record.get("principal", 0))

    total_due = float(
        latest_record.get(
            "total_due",
            latest_record.get(
                "amount_due",
                0
            )
        )
    )
    
    paid = filtered_loans["amount_paid"].fillna(0).astype(float).sum()
    
    bal = float(
        latest_record.get(
            "balance",
            0
        )
    )
    
    i = max(
        total_due - p,
        0
    )
    
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Principal Allocation", f"UGX {p:,.0f}", delta="Disbursement Base", delta_color="off")
    m2.metric("Interest Cost Accrued", f"UGX {i:,.0f}", delta=f"{(i/total_due*100):.1f}% Markup" if total_due > 0 else None, delta_color="normal")
    m3.metric("Total Paid to Date", f"UGX {paid:,.0f}", delta=f"{paid/total_due:.1%} Covered" if total_due > 0 else None, delta_color="normal")
    m4.metric("Current Balance Owed", f"UGX {bal:,.0f}", delta=f"Outstanding Position", delta_color="inverse")

    # ==============================
    # 📜 CORRECTED TRANSACTION HISTORY (LEDGER RUNTIME)
    # ==============================
    ledger_data = []

    for _, row in filtered_loans.iterrows():
    
        ledger_data.append({
    
            "Cycle":
                row.get("cycle", ""),
    
            "Date":
                str(
                    row.get(
                        "start_date",
                        row.get(
                            "created_at",
                            ""
                        )
                    )
                )[:10],
    
            "Opening Balance":
                float(
                    row.get(
                        "principal",
                        row.get(
                            "opening_balance",
                            0
                        )
                    )
                ),
    
            "Amount Due":
                float(
                    row.get(
                        "total_due",
                        row.get(
                            "amount_due",
                            0
                        )
                    )
                ),
    
            "Amount Paid":
                float(
                    row.get(
                        "amount_paid",
                        0
                    )
                ),
    
            "Balance":
                float(
                    row.get(
                        "balance",
                        0
                    )
                ),
    
            "Status":
                row.get(
                    "status",
                    ""
                )
        })
    
    ledger_df = pd.DataFrame(ledger_data)
    
    st.dataframe(
        ledger_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Opening Balance":
                st.column_config.NumberColumn(
                    format="%,.0f"
                ),
    
            "Amount Due":
                st.column_config.NumberColumn(
                    format="%,.0f"
                ),
    
            "Amount Paid":
                st.column_config.NumberColumn(
                    format="%,.0f"
                ),
    
            "Balance":
                st.column_config.NumberColumn(
                    format="%,.0f"
                ),
        }
    )

    st.markdown("---")

    # ==============================
    # 📄 PREMIUM DOWNLOAD SECTION
    # ==============================
    st.markdown(f"""
        <div style="border: 1px solid {brand_color}33; padding: 1.5rem; border-radius: 12px; background-color: {brand_color}08; margin-bottom:15px;">
            <p style="font-family: 'Inter', sans-serif; font-weight: 600; margin-bottom: 5px; color:{brand_color};">Export Premium Statement Artifact</p>
            <p style="font-family: 'Inter', sans-serif; font-size: 0.88rem; color: #4B5563; margin-bottom: 0px;">
                Generates audit-ready customer ledgers complete with formal registration letterhead blocks.
            </p>
        </div>
    """, unsafe_allow_html=True)

    if st.button("✨ Compile Formal PDF Statement", use_container_width=True):
        client_name = filtered_loans.iloc[0].get("borrower", "Unknown Profile")
        client_loans = loans_df[loans_df["borrower"] == client_name]

        with st.spinner("Rendering Document Architecture..."):
            pdf_output = generate_pdf_statement(client_name, client_loans, payments_df)

        st.download_button(
            label=f"⬇️ Download Verified Statement Document (.pdf)",
            data=pdf_output,
            file_name=f"Statement_{client_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf",
            mime="application/pdf",
            use_container_width=True
        )

def show_settings():
    """
    Manages tenant identity and UI branding.
    Only displays when the 'Settings' page is selected.
    """
    # ==============================
    # 🔐 TENANT SAFETY LAYER (HARD GUARD)
    # ==============================
    tenant_id = st.session_state.get("tenant_id")

    if not tenant_id:
        st.warning("⚠️ No Active tenant detected. Please log in.")
        return

    # ==============================
    # 1. FETCH TENANT DATA (SAFE + HARDENED)
    # ==============================
    try:
        # Fetching the business profile specifically for this tenant
        tenant_resp = supabase.table("tenants").select("*").eq("id", tenant_id).execute()

        if not tenant_resp.data:
            st.error("❌ Business profile not found.")
            return

        active_company = tenant_resp.data[0]

    except Exception as e:
        st.error(f"❌ Connection Error: {e}")
        return

    # ==============================
    # BRANDING FALLBACK SAFETY
    # ==============================
    brand_color = st.session_state.get(
        "theme_color", 
        active_company.get("brand_color", "#2B3F87")
    )

    st.markdown(
        f"<h2 style='color: {brand_color};'>⚙️ Portal Settings & Branding</h2>",
        unsafe_allow_html=True
    )

    # --- BUSINESS IDENTITY SECTION ---
    st.subheader("🏢 Business Identity")

    col1, col2 = st.columns([2, 1])

    with col1:
        st.markdown(f"**Current Business name:** {active_company.get('name', 'Unknown')}")

        new_color = st.color_picker(
            "🎨 Change Brand Color",
            active_company.get('brand_color', '#2B3F87'),
            key="settings_color_picker"
        )

        st.markdown("**Preview:**")
        st.markdown(
            f"""
            <div style='padding:15px; 
                        background-color:{new_color}; 
                        color:white; 
                        border-radius:10px; 
                        text-align:center; 
                        font-weight:bold;'>
                Brand Color Preview
            </div>
            """, 
            unsafe_allow_html=True
        )

    with col2:
        st.markdown("**Company Logo:**")
        
        logo_url = active_company.get("logo_url")

        # ==============================
        # LOGO DISPLAY SAFETY (CACHE BUSTING)
        # ==============================
        if logo_url:
            try:
                # Append timestamp to URL to force browser refresh on logo update
                logo_display_url = f"{logo_url}?t={int(time.time())}"
                st.image(logo_display_url, use_container_width=True, caption="Current Logo")
            except Exception:
                st.caption("⚠️ Logo could not be loaded.")
        else:
            st.caption("No logo uploaded yet.")

        logo_file = st.file_uploader("Upload New Logo (PNG/JPG)", type=["png", "jpg", "jpeg"])

    st.divider()

    # --- SAVE ACTION ---
    if st.button("💾 Save Branding Changes", use_container_width=True):
        
        updated_data = {"brand_color": new_color}

        # ==============================
        # LOGO UPLOAD SAFETY (STORAGE BUCKET FIXED)
        # ==============================
        if logo_file:
            try:
                bucket_name = "company-logos"
                file_path = f"logos/{active_company.get('id')}_logo.png"
                file_bytes = logo_file.getvalue()

                # Robust Upsert Logic: Try updating first; if it doesn't exist, execute an upload.
                try:
                    supabase.storage.from_(bucket_name).update(
                        path=file_path,
                        file=file_bytes,
                        file_options={"content-type": "image/png"}
                    )
                except Exception:
                    # Fallback to fresh file creation if the path didn't exist yet
                    supabase.storage.from_(bucket_name).upload(
                        path=file_path,
                        file=file_bytes,
                        file_options={"content-type": "image/png"}
                    )

                # Generate public URL for database storage
                public_url = supabase.storage.from_(bucket_name).get_public_url(file_path)
                updated_data["logo_url"] = public_url

            except Exception as e:
                st.error(f"❌ Storage Error: {str(e)}")
                st.stop()

        # ==============================
        # DATABASE UPDATE (PERSISTENCE)
        # ==============================
        try:
            supabase.table("tenants").update(updated_data).eq("id", active_company.get("id")).execute()
            
            # Immediately update session state for real-time UI feel
            st.session_state["theme_color"] = new_color
            if "logo_url" in updated_data:
                st.session_state["logo_url"] = updated_data["logo_url"]

            st.success("✅ Branding updated successfully!")
            time.sleep(1)
            st.rerun()

        except Exception as e:
            st.error(f"❌ Database Error: {str(e)}")



# ==========================================
# FINAL APP ROUTER (REACTIVE & STABLE)
# ==========================================

if __name__ == "__main__":

    # 1. Initialize Default State
    if "logged_in" not in st.session_state:
        st.session_state["logged_in"] = False

    if "view" not in st.session_state:
        st.session_state["view"] = "login"

    # 2. 🔐 AUTH FLOW
    if not st.session_state.get("logged_in"):
        st.session_state['theme_color'] = "#1E3A8A"
        apply_master_theme()
        run_auth_ui(supabase)
        # Note: run_auth_ui handles its own views. 
        # We don't want the rest of the script to run if not logged in.
    
    # 3. 🚀 MAIN APP (Only runs if logged_in is True)
    else:
        try:
            check_session_timeout()

            # Sidebar (Get the selected page)
            raw_page = render_sidebar()
            
            # Theme
            apply_master_theme()

            # 🔥 Clean the page string to ensure matching works perfectly
            # This handles any accidental spaces or casing issues
            page = str(raw_page).strip()

            # 4. 🗺️ NAVIGATION ROUTER
            if page == "Overview":
                show_dashboard_view()
                
            elif page == "loans":
                show_loans()
                
            elif page == "borrowers":
                show_borrowers()
                
            elif page == "Collateral":
                show_collateral()
                
            elif page == "Calendar":
                show_calendar()
                
            elif page == "Ledger":
                show_ledger()
                
            elif page == "Payments":
                show_payments(supabase)
                
            elif page == "Expenses":
                show_expenses()
                
            elif page == "Overdue Tracker":
                show_overdue_tracker()
                
            elif page == "Payroll":
                show_payroll()
                
            elif page == "Reports":
                show_reports()
                
            elif page == "Settings":
                show_settings()

            else:
                # If it falls through here, we show what exactly was received
                st.info(f"Module '{page}' is coming online soon.")
                # Debugging help:
                # st.write(f"DEBUG: Sidebar returned '{page}'")

        except Exception as e:
            st.error(f"🚨 Application Error: {e}")
            if st.button("Clear Cache & Retry"):
                st.cache_data.clear()
                st.rerun()
